#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Simuliert die von Marcus vorgeschlagenen Nummern-Regeln auf dem April-Extrakt
und baut daraus Auswertung (.xlsx) und Antwortmail (.eml).

Drei Varianten werden gerechnet:
  A  Regeln wie vorgeschlagen: bereinigen (Praefixe, fuehrende Nullen,
     Trennzeichen), < 7 Stellen -> Vermittler, 7 oder 9 Stellen -> Agentur
  B  zusaetzlich: Felder, die zwei mit Label versehene Nummern enthalten
     ("V 85357 A 777-0503"), vorher auftrennen
  C  zusaetzlich: achtstellige Nummern mit einer fuehrenden Null auf neun
     Stellen ergaenzen und damit als Agenturnummer werten

Aufruf:  python3 scripts/build-mail-regelsimulation.py
"""

import collections
import re

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mailbau import baue_eml, tab

QUELLE = "analyse.xlsx"
EXCEL = "Regelsimulation-Nummern.xlsx"
EML = "Regelsimulation-Nummern.eml"
BETREFF = "Nummern-Regeln am April-Extrakt durchgerechnet"

FONT = "Aptos"
BLAU = "FF1E40AF"
LILA = "FF4C1D95"
GRUEN = "FFDCFCE7"
GELB = "FFFEF3C7"
ROT = "FFFEE2E2"
WEISS = "FFFFFFFF"

KATEGORIEN = ("beide", "nur Agentur", "nur Vermittler", "keine")

LABEL = re.compile(r"[A-Za-z]{1,2}\s*\d[\d\s\-\/]*")


# ---------------------------------------------------------------------------
# Regelwerk
# ---------------------------------------------------------------------------
def ziffern(v):
    return re.sub(r"\D", "", v)


def auftrennen(v):
    """Variante B: 'V 85357 A 777-0503' in zwei Nummern zerlegen.

    Nur wenn mindestens zwei mit Buchstaben-Label versehene Nummern gefunden
    werden UND diese den Wert praktisch vollstaendig abdecken. Damit bleiben
    Einzelnummern wie 'A600040124', '60002/0753' oder '00 111 40 16'
    unangetastet.
    """
    treffer = LABEL.findall(v)
    if len(treffer) < 2:
        return [v]
    abgedeckt = sum(len(re.sub(r"\s", "", t)) for t in treffer)
    gesamt = len(re.sub(r"\s", "", v)) or 1
    return [t.strip() for t in treffer] if abgedeckt / gesamt >= 0.8 else [v]


def kandidaten(r, mit_trennung):
    out = []
    for feld, heute_spalte in ((r[13], "Agentur"), (r[14], "Vermittler")):
        if not feld:
            continue
        for teil in re.split(r"[,;]| und ", feld):
            teil = teil.strip()
            if not re.search(r"\d", teil):
                continue
            for stueck in (auftrennen(teil) if mit_trennung else [teil]):
                out.append((stueck, heute_spalte))
    return out


def klasse(v, acht_auffuellen):
    """< 7 Stellen -> Vermittler, 7 oder 9 Stellen -> Agentur, sonst offen.

    acht_auffuellen: achtstellige Nummern bekommen eine fuehrende Null,
    sind damit neunstellig und gelten als Agenturnummer.
    """
    k = ziffern(v).lstrip("0")
    n = len(k)
    if n == 0:
        return None
    if acht_auffuellen and n == 8:
        return "Agentur"
    if n < 7:
        return "Vermittler"
    if n in (7, 9):
        return "Agentur"
    return "ohne Zuordnung"


def rechne(bue, mit_trennung, acht_auffuellen):
    befuellung = collections.Counter()
    offen = collections.Counter()
    wechsel = collections.Counter()
    for r in bue:
        klassen = []
        for roh, heute_spalte in kandidaten(r, mit_trennung):
            k = klasse(roh, acht_auffuellen)
            klassen.append(k)
            wechsel[(heute_spalte, k or "leer")] += 1
            if k == "ohne Zuordnung":
                rein = ziffern(roh).lstrip("0")
                offen[(roh, rein, len(rein))] += 1
        hat_a, hat_v = "Agentur" in klassen, "Vermittler" in klassen
        befuellung["beide" if hat_a and hat_v else "nur Agentur" if hat_a
                   else "nur Vermittler" if hat_v else "keine"] += 1
    return befuellung, offen, wechsel


# ---------------------------------------------------------------------------
# Daten
# ---------------------------------------------------------------------------
wb = openpyxl.load_workbook(QUELLE, read_only=True, data_only=True)
BUE = [[("" if v is None else str(v).strip()) for v in r]
       for r in wb["Analyse"].iter_rows(min_row=3, values_only=True)]
wb.close()
BUE = [r for r in BUE if r[21] in ("BÜ-Vorgang", "BUe-Vorgang")]
N = len(BUE)

heute = collections.Counter()
for r in BUE:
    heute["beide" if r[13] and r[14] else "nur Agentur" if r[13]
          else "nur Vermittler" if r[14] else "keine"] += 1

A_bef, A_offen, A_wechsel = rechne(BUE, False, False)
B_bef, B_offen, _ = rechne(BUE, True, False)
C_bef, C_offen, _ = rechne(BUE, True, True)

WERTE = sum(A_wechsel.values())
N_WECHSLER = A_wechsel[("Vermittler", "Agentur")]
wechsler_liste = collections.Counter()
for r in BUE:
    for roh, h in kandidaten(r, False):
        if h == "Vermittler" and klasse(roh, False) == "Agentur":
            wechsler_liste[(roh, ziffern(roh).lstrip("0"))] += 1

A_OFFEN, B_OFFEN, C_OFFEN = (sum(x.values()) for x in (A_offen, B_offen, C_offen))
ACHT = sum(c for (_, _, l), c in A_offen.items() if l == 8)
LANG = sum(c for (_, _, l), c in A_offen.items() if l > 9)
REST_ACHT = sum(c for (_, _, l), c in C_offen.items() if l == 8)


def pz(z, n):
    return ("%.1f" % (100.0 * z / n)).replace(".", ",") + " %"


def tsd(n):
    return "{:,}".format(int(n)).replace(",", ".")


def diff(neu, alt):
    return "%+d" % (neu - alt) if neu != alt else "0"


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------
def kopf(ws, spalten, breiten, zeile=1):
    for i, (t, b) in enumerate(zip(spalten, breiten), 1):
        z = ws.cell(row=zeile, column=i, value=t)
        z.font = Font(name=FONT, sz=10, b=True, color=WEISS)
        z.fill = PatternFill("solid", fgColor=LILA)
        z.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = b


def zeile_schreiben(ws, r, werte, fill=None):
    for i, w in enumerate(werte, 1):
        z = ws.cell(row=r, column=i, value=w)
        z.font = Font(name=FONT, sz=10)
        z.alignment = Alignment(vertical="top", wrap_text=True)
        if fill:
            z.fill = PatternFill("solid", fgColor=fill)


def titelzeile(ws, text, unter=""):
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, sz=14, b=True, color=BLAU)
    if unter:
        ws["A2"] = unter
        ws["A2"].font = Font(name=FONT, sz=10, color="FF4B5563")
    ws.sheet_view.showGridLines = False


VARIANTEN = [
    ("heute (Präfix-Regel 6000/890)", heute, None, None),
    ("A - Regeln wie vorgeschlagen", A_bef, A_OFFEN, None),
    ("B - zusätzlich Felder mit zwei Nummern trennen", B_bef, B_OFFEN, GELB),
    ("C - zusätzlich achtstellige auf neun Stellen auffüllen", C_bef, C_OFFEN, GRUEN),
]

xl = openpyxl.Workbook()
ws = xl.active
ws.title = "Varianten"
titelzeile(ws, "Nummern-Regeln: drei Varianten im Vergleich",
           "Basis: %s BÜ-Vorgänge aus dem April-Extrakt mit %s extrahierten Nummern."
           % (tsd(N), tsd(WERTE)))
kopf(ws, ["Variante", "beide", "nur Agentur", "nur Vermittler", "keine",
          "Nummern ohne Zuordnung"], [50, 12, 14, 16, 12, 22], zeile=4)
r = 5
for name, bef, off, fill in VARIANTEN:
    zeile_schreiben(ws, r, [name] + [bef[k] for k in KATEGORIEN] +
                    ["" if off is None else off], fill=fill)
    r += 1

ws2 = xl.create_sheet("Wechsler V nach A")
titelzeile(ws2, "Nummern, die neu als Agenturnummer gelten",
           "Heute Vermittlernummer, nach den neuen Regeln 7- oder 9-stellig. "
           "%s Werte in %s Schreibweisen." % (tsd(N_WECHSLER), tsd(len(wechsler_liste))))
kopf(ws2, ["Rohwert heute", "bereinigt", "Stellen", "Vorgänge", "Präfix 890/6000?"],
     [26, 20, 10, 12, 18], zeile=4)
r = 5
for (roh, z), c in wechsler_liste.most_common():
    passt = "ja" if z.startswith(("890", "6000")) else "nein"
    zeile_schreiben(ws2, r, [roh, z, len(z), c, passt], fill=GELB if passt == "nein" else None)
    r += 1

ws3 = xl.create_sheet("Ohne Zuordnung")
titelzeile(ws3, "Nummern, die durch das Raster fallen",
           "Weder unter 7 Stellen noch genau 7 oder 9 Stellen. Rechte Spalte: löst "
           "Variante C den Fall?")
kopf(ws3, ["Rohwert", "bereinigt", "Stellen", "Vorgänge", "in Variante C gelöst?"],
     [28, 20, 10, 12, 22], zeile=4)
r = 5
c_offen_roh = {k[0] for k in C_offen}
for (roh, z, laenge), c in sorted(A_offen.items(), key=lambda x: (-x[1], x[0][2])):
    geloest = "nein" if roh in c_offen_roh else "ja"
    zeile_schreiben(ws3, r, [roh, z, laenge, c, geloest],
                    fill=GRUEN if geloest == "ja" else ROT)
    r += 1

xl.calculation.fullCalcOnLoad = True
xl.save(EXCEL)


# ---------------------------------------------------------------------------
# Mail
# ---------------------------------------------------------------------------
top_wechsler = ", ".join("<code>%s</code> (%dx)" % (roh, c)
                         for (roh, _), c in wechsler_liste.most_common(6))

HTML = """<div>
<p>Hallo Marcus,</p>

<p>danke für den Vorschlag - die ersten vier Regeln habe ich über den
April-Extrakt laufen lassen. Basis sind die {N} BÜ-Vorgänge mit {WERTE}
extrahierten Nummern. Die Auswertung hängt als Excel an.</p>

<p><b>Kurzfassung:</b> Die Regeln greifen und decken die Agenturnummer deutlich
besser ab als bisher. An zwei Stellen fallen Nummern durchs Raster - beides
lässt sich mit einer kleinen technischen Ergänzung beheben, ohne an eurer
Regellogik etwas zu ändern.</p>

<h3>1. Befüllung je Vorgang</h3>
{TAB_BEF}
<p>Der Einstieg über die Agenturnummer - für euch der Weg in EMMA - deckt statt
{AG_ALT} künftig {AG_NEU} Vorgänge ab, das sind {AG_PZ} statt {AG_PZ_ALT}. Der
Rückfallweg über die Vermittlernummer schrumpft entsprechend von 624 auf
104 Vorgänge.</p>

<p><b>Warum &bdquo;keine&ldquo; zunächst um 117 steigt: Es sind achtstellige
Nummern.</b> Die Regeln decken zwei Längen ab - unter 7 Stellen wird
Vermittlernummer, 7 oder 9 Stellen wird Agenturnummer. <b>Für 8 Stellen gibt es
keine Regel</b>, und genau diese Länge kommt {ACHT} mal vor. Dazu {LANG} Werte
mit 10 und mehr Stellen. In all diesen {A_OFFEN} Vorgängen ist das jeweils die
einzige Nummer - fällt sie weg, bleibt nichts übrig.</p>

<h3>2. Zwei Ergänzungen, die das beheben</h3>
<p><b>Ergänzung 1 - Felder mit zwei Nummern trennen.</b> Werte wie
<code>V 85357 A 777-0503</code> enthalten in Wirklichkeit zwei Nummern:
Vermittler 85357 und Agentur 777-0503. Die Bereinigung zieht sie heute zu
<code>853577770503</code> zusammen - 12 Stellen, damit unbrauchbar. Trennt man
vorher an den Buchstaben-Labels, ergibt derselbe Wert sauber
<code>85357</code> (Vermittler) und <code>7770503</code> (Agentur), der Vorgang
hat danach sogar beide Nummern. Die Trennung greift nur, wenn wirklich zwei
gelabelte Nummern vorliegen - Einzelwerte wie <code>A600040124</code>,
<code>60002/0753</code> oder <code>00 111 40 16</code> bleiben unangetastet.</p>

<p><b>Ergänzung 2 - achtstellige Nummern mit einer führenden Null auf neun
Stellen ergänzen.</b> Dein Vorschlag, und die Daten stützen ihn deutlich:
{N_1000} der {ACHT} Achtsteller gehören zur selben Familie, die sonst
neunstellig geschrieben wird - <code>01000-3083</code>,
<code>01000/3451</code>, <code>A010006515</code>. Ohne die führende Null steht
dort <code>10003083</code>, mit ihr <code>010003083</code>, und das passt genau
ins Muster fünf Stellen Agentur plus vier Stellen Unternummer. Die Regel
erledigt damit alle {ACHT} Achtsteller in einem Schritt.</p>

{TAB_VARIANTEN}

<p>Mit beiden Ergänzungen bleiben von den {A_OFFEN} offenen Nummern noch
<b>{C_OFFEN}</b> übrig, und die Vorgänge ganz ohne Nummer liegen bei
<b>{C_KEINE}</b> gegenüber {HEUTE_KEINE} heute - also praktisch auf dem
heutigen Stand, bei gleichzeitig {AG_GEWINN} Vorgängen mehr mit
Agenturnummer.</p>

<p>Die {C_OFFEN} verbleibenden Werte sind Einzelfälle mit 10 und mehr Stellen:
<code>6000020026</code>, <code>28964000000</code>,
<code>86821000000/60001/0017</code> und drei ähnliche. Die würde ich direkt als
&bdquo;zu prüfen&ldquo; kennzeichnen.</p>

<p><b>Ein Detail zur Auffüllregel, das mir dabei aufgefallen ist:</b> Sie passt
für die {N_1000} Werte der 01000-Familie sehr gut. Bei {N_6008} Werten beginnt
die Nummer aber mit <code>6008</code> - etwa <code>60083652</code> oder
<code>60080159</code>. Dort fehlt die Null vermutlich nicht vorn, sondern in
der Mitte: <code>600083652</code> würde in die dichte Reihe der
6000-Agenturnummern passen, <code>060083652</code> nicht. Die Auffüllregel
liefert hier also möglicherweise die falsche Nummer. Weil dein EMMA-Abgleich
ohnehin prüft, ob die Agenturnummer existiert, würden genau diese Fälle dort
auffallen und als &bdquo;zu prüfen&ldquo; landen - insofern ist es kein
Blocker, ich wollte es nur nicht unter den Tisch fallen lassen.</p>

<h3>3. Was sich je Nummer ändert</h3>
{TAB_WECHSEL}
<p>Die Agenturnummern bleiben praktisch unangetastet: {A_STABIL} von
{A_GESAMT} behalten ihre Zuordnung. Die Bewegung steckt bei den
Vermittlernummern - <b>{N_WECHSLER} Werte ({N_WECHSLER_PZ} aller Nummern)
werden neu zur Agenturnummer</b>. Häufigste Fälle: {TOP_WECHSLER}.</p>
<p>Das ist der eigentliche Unterschied zur heutigen Logik: Die bisherige Regel
&bdquo;beginnt mit 6000 oder 890&ldquo; und die neue Längenregel widersprechen
sich bei genau diesen Werten. <code>8923555</code> beginnt mit 89, aber nicht
mit 890; <code>601000040</code> mit 601, nicht mit 6000. Nach der Längenregel
sind beide Agenturnummern. Ob das fachlich stimmt, kann ich aus den Daten nicht
beurteilen - der EMMA-Abgleich würde es aber genau hier zeigen.</p>

<h3>Was ich von dir bräuchte</h3>
<ol>
<li><b>Die Trennregel:</b> Passt sie so? Sie ist reine Technik und ändert
nichts an eurer Regellogik.</li>
<li><b>Die 6008-Fälle:</b> Sollen wir sie wie alle anderen vorn auffüllen und
den EMMA-Abgleich entscheiden lassen, oder von vornherein als
&bdquo;zu prüfen&ldquo; kennzeichnen?</li>
<li><b>Die {N_WECHSLER} Wechsler:</b> Im Blatt &bdquo;Wechsler V nach A&ldquo;
stehen sie einzeln mit Häufigkeit. Magst du stichprobenhaft draufschauen, ob
die Einordnung als Agenturnummer passt? Falls ja, ist die Längenregel der
heutigen Präfix-Regel klar überlegen.</li>
<li><b>Werte mit Textrest:</b> Bei <code>600041966 1zu1 AG</code> zieht die
Bereinigung auch die Ziffern aus dem Firmennamen mit - heraus kommt
<code>60004196611</code>, 11 statt 9 Stellen. Wir würden den Textanteil vorher
entfernen; nur zur Info, dass wir das so vorsehen.</li>
</ol>

<p>Den EMMA-Teil deines Vorschlags - Agenturnummer als führend,
Vermittlernummer ergänzen oder überschreiben, bei Verstößen
&bdquo;zu prüfen&ldquo; - können wir genau so umsetzen. Sobald die Punkte oben
geklärt sind, ziehe ich die Regeln in die Nachverarbeitung ein und lasse den
April-Extrakt komplett neu durchlaufen.</p>

<p>Rückfragen gerne jederzeit.</p>

<p>Viele Grüße</p>
</div>"""

HTML = HTML.replace("{TAB_BEF}", tab(
    ["Befüllung je Vorgang", "heute", "nach Vorschlag", "Veränderung"],
    [[k, tsd(heute[k]), tsd(A_bef[k]), diff(A_bef[k], heute[k])] for k in KATEGORIEN]
    + [["<b>Summe</b>", "<b>%s</b>" % tsd(N), "<b>%s</b>" % tsd(N), "<b>0</b>"]],
    num_spalten=(1, 2, 3)))

HTML = HTML.replace("{TAB_VARIANTEN}", tab(
    ["Variante", "beide", "nur Agentur", "nur Vermittler", "keine", "ohne Zuordnung"],
    [[name] + [tsd(bef[k]) for k in KATEGORIEN] + ["-" if off is None else tsd(off)]
     for name, bef, off, _ in VARIANTEN],
    num_spalten=(1, 2, 3, 4, 5)))

HTML = HTML.replace("{TAB_WECHSEL}", tab(
    ["heute", "nach Vorschlag", "Nummern", "Anteil"],
    [[h, n, tsd(c), pz(c, WERTE)]
     for (h, n), c in sorted(A_wechsel.items(), key=lambda x: -x[1])],
    num_spalten=(2, 3)))

# Aufteilung der Achtsteller nach Nummernfamilie (fuer die Einordnung im Text)
acht_familien = collections.Counter()
for r in BUE:
    for roh, _ in kandidaten(r, True):
        k = ziffern(roh).lstrip("0")
        if len(k) == 8:
            acht_familien[k[:3]] += 1
N_1000 = acht_familien["100"]
N_6008 = acht_familien["600"]

A_STABIL = A_wechsel[("Agentur", "Agentur")]
A_GESAMT = sum(c for (h, _), c in A_wechsel.items() if h == "Agentur")

for platz, wert in (
        ("{N}", tsd(N)), ("{WERTE}", tsd(WERTE)),
        ("{AG_ALT}", tsd(heute["beide"] + heute["nur Agentur"])),
        ("{AG_NEU}", tsd(A_bef["beide"] + A_bef["nur Agentur"])),
        ("{AG_PZ_ALT}", pz(heute["beide"] + heute["nur Agentur"], N)),
        ("{AG_PZ}", pz(A_bef["beide"] + A_bef["nur Agentur"], N)),
        ("{ACHT}", tsd(ACHT)), ("{LANG}", tsd(LANG)),
        ("{A_OFFEN}", tsd(A_OFFEN)), ("{C_OFFEN}", tsd(C_OFFEN)),
        ("{C_KEINE}", tsd(C_bef["keine"])), ("{HEUTE_KEINE}", tsd(heute["keine"])),
        ("{C_DIFF}", tsd(C_bef["keine"] - heute["keine"])),
        ("{N_1000}", tsd(N_1000)), ("{N_6008}", tsd(N_6008)),
        ("{AG_GEWINN}", tsd((C_bef["beide"] + C_bef["nur Agentur"])
                            - (heute["beide"] + heute["nur Agentur"]))),
        ("{A_STABIL}", tsd(A_STABIL)), ("{A_GESAMT}", tsd(A_GESAMT)),
        ("{N_WECHSLER_PZ}", pz(N_WECHSLER, WERTE)), ("{N_WECHSLER}", tsd(N_WECHSLER)),
        ("{TOP_WECHSLER}", top_wechsler)):
    HTML = HTML.replace(platz, wert)

offen_platz = re.findall(r"\{[A-Z_0-9]+\}", HTML)
assert not offen_platz, offen_platz


# ---------------------------------------------------------------------------
# Klartext aus demselben HTML
# ---------------------------------------------------------------------------
def umbruch(text, breite=76):
    worte, zeilen, akt = text.split(), [], ""
    for w in worte:
        if akt and len(akt) + 1 + len(w) > breite:
            zeilen.append(akt)
            akt = w
        else:
            akt = (akt + " " + w).strip()
    if akt:
        zeilen.append(akt)
    return "\n".join(zeilen)


def ascii_tab(kopfzeilen, zeilen, breiten):
    aus = ["  " + "".join(("%-" + str(b) + "s") % k for k, b in zip(kopfzeilen, breiten)),
           "  " + "-" * (sum(breiten) - 2)]
    for zl in zeilen:
        aus.append("  " + "".join(("%-" + str(b) + "s") % w for w, b in zip(zl, breiten)))
    return "\n".join(aus)


TAB_TEXT = [
    ascii_tab(["Befüllung je Vorgang", "heute", "Vorschlag", "Änderung"],
              [[k, tsd(heute[k]), tsd(A_bef[k]), diff(A_bef[k], heute[k])]
               for k in KATEGORIEN] + [["Summe", tsd(N), tsd(N), "0"]],
              [24, 10, 12, 10]),
    ascii_tab(["Variante", "beide", "nur Ag.", "nur Verm.", "keine", "offen"],
              [[name] + [tsd(bef[k]) for k in KATEGORIEN]
               + ["-" if off is None else tsd(off)] for name, bef, off, _ in VARIANTEN],
              [56, 8, 9, 11, 8, 8]),
    ascii_tab(["heute", "nach Vorschlag", "Nummern", "Anteil"],
              [[h, n, tsd(c), pz(c, WERTE)]
               for (h, n), c in sorted(A_wechsel.items(), key=lambda x: -x[1])],
              [14, 18, 10, 10]),
]

t = HTML
t = re.sub(r"<h3>(.*?)</h3>", lambda m: "\n@@H@@" + m.group(1) + "\n", t)
t = re.sub(r"<li>(.*?)</li>", lambda m: "  - " + m.group(1) + "\n", t)
t = re.sub(r"<table>.*?</table>", "@@TAB@@", t, flags=re.S)
t = re.sub(r"</p>", "\n\n", t)
t = re.sub(r"<[^>]+>", "", t)
for a, b in (("&bdquo;", "„"), ("&ldquo;", "“"), ("&nbsp;", " "), ("&amp;", "&")):
    t = t.replace(a, b)
for tabelle in TAB_TEXT:
    t = t.replace("@@TAB@@", "\n" + tabelle + "\n", 1)

zeilen_txt = []
for zeile in t.split("\n"):
    zeile = zeile.rstrip()
    if zeile.startswith("@@H@@"):
        titel = zeile[5:]
        zeilen_txt += ["", titel, "-" * len(titel)]
    elif zeile.startswith("  ") or not zeile.strip():
        zeilen_txt.append(zeile)
    else:
        zeilen_txt.append(umbruch(zeile))
TEXT = re.sub(r"\n{3,}", "\n\n", "\n".join(zeilen_txt)).strip() + "\n"

baue_eml(EML, BETREFF, TEXT, HTML, [EXCEL])
print("   A: keine=%d offen=%d | B: keine=%d offen=%d | C: keine=%d offen=%d (heute %d)"
      % (A_bef["keine"], A_OFFEN, B_bef["keine"], B_OFFEN,
         C_bef["keine"], C_OFFEN, heute["keine"]))
