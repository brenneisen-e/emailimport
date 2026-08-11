#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Baut den Feldkatalog des KI-Deckblatts als Excel-Datei.

Quelle ist datenfelder.py - dasselbe Modul, aus dem auch die Mail erzeugt
wird. Drei Blaetter:
  * "Datenfelder"          - alle Felder mit Beschreibung, Ausprägungen,
                             Beispiel, Anpassung und Status
  * "Prüfung Unterlagen"   - Ausprägungen des abgeleiteten Prüffelds
  * "Offene Punkte"        - was noch fachlich zu klären ist

Aufruf:  python3 scripts/build-datenfelder-xlsx.py [ziel.xlsx]
"""

import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from datenfelder import ABSCHNITTE, BESONDERHEITEN, OFFEN, PRUEFUNG, rein

FONT = "Aptos"
BLAU = "FF1E40AF"
LILA = "FF4C1D95"
GRAU = "FFF3F4F6"
GRUEN = "FFDCFCE7"
GELB = "FFFEF3C7"
WEISS = "FFFFFFFF"

THIN = Side(style="thin", color="FFD1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_TEXT = {"neu": "NEU", "offen": "OFFEN", None: ""}
STATUS_FARBE = {"neu": GRUEN, "offen": GELB, None: None}


def kopfzeile(ws, spalten, breiten, zeile=1):
    for i, (titel, breite) in enumerate(zip(spalten, breiten), 1):
        z = ws.cell(row=zeile, column=i, value=titel)
        z.font = Font(name=FONT, sz=10, b=True, color=WEISS)
        z.fill = PatternFill("solid", fgColor=LILA)
        z.alignment = Alignment(vertical="center", wrap_text=True)
        z.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = breite
    ws.row_dimensions[zeile].height = 22


def schreibe(ws, zeile, werte, fill=None, bold_erste=False):
    for i, wert in enumerate(werte, 1):
        z = ws.cell(row=zeile, column=i, value=wert)
        z.font = Font(name=FONT, sz=10, b=(bold_erste and i == 1))
        z.alignment = Alignment(vertical="top", wrap_text=True)
        z.border = BORDER
        if fill:
            z.fill = PatternFill("solid", fgColor=fill)


def hoehe(ws, zeile, texte, breiten):
    """Zeilenhöhe grob aus dem laengsten Zellinhalt je Spaltenbreite."""
    zeilen = max(1, max((len(t) // max(10, int(b * 1.05)) + 1)
                        for t, b in zip(texte, breiten) if isinstance(t, str)) if texte else 1)
    ws.row_dimensions[zeile].height = 14 * min(zeilen, 8)


def baue_xlsx(ziel):
    wb = openpyxl.Workbook()

    # ---- Blatt 1: Datenfelder -------------------------------------------
    ws = wb.active
    ws.title = "Datenfelder"
    ws["A1"] = "Datenfelder KI-Deckblatt - Zielbild"
    ws["A1"].font = Font(name=FONT, sz=14, b=True, color=BLAU)
    ws["A2"] = ("Alle Felder, die auf dem KI-Deckblatt erscheinen, inklusive der Anpassungen "
                "aus dem Review 260807 Anpassung Deckblatt_V1.pdf. Zusatzspalten, die es nur "
                "in der Excel-Auswertung gibt, sind nicht aufgeführt.")
    ws["A2"].font = Font(name=FONT, sz=10, color="FF4B5563")
    ws["A3"] = ("Schreibweise: In allen Adressfeldern wird \"Straße\" grundsätzlich als "
                "\"Str.\" ausgegeben. Status NEU = neues, getrenntes oder geändertes Feld; "
                "OFFEN = fachliche Vorgabe fehlt noch.")
    ws["A3"].font = Font(name=FONT, sz=10, color="FF4B5563")

    spalten = ["Abschnitt", "Feld", "Beschreibung", "Mögliche Ausprägungen / Format",
               "Beispiel", "Anpassung", "Status"]
    breiten = [26, 24, 46, 46, 30, 40, 9]
    kopfzeile(ws, spalten, breiten, zeile=5)

    zeile = 6
    for titel, felder in ABSCHNITTE:
        for feld, beschr, auspr, beisp, anp, mark in felder:
            werte = [rein(titel), rein(feld), rein(beschr), rein(auspr), rein(beisp),
                     rein(anp), STATUS_TEXT[mark]]
            schreibe(ws, zeile, werte, fill=STATUS_FARBE[mark], bold_erste=False)
            ws.cell(row=zeile, column=2).font = Font(name=FONT, sz=10, b=True)
            hoehe(ws, zeile, werte, breiten)
            zeile += 1

    ws.freeze_panes = "C6"
    ws.auto_filter.ref = "A5:%s%d" % (get_column_letter(len(spalten)), zeile - 1)
    ws.sheet_view.showGridLines = False

    # ---- Blatt 2: Prüfung Unterlagen ------------------------------------
    ws2 = wb.create_sheet("Prüfung Unterlagen")
    ws2["A1"] = "Ausprägungen des Feldes \"Prüfung Unterlagen\""
    ws2["A1"].font = Font(name=FONT, sz=14, b=True, color=BLAU)
    ws2["A2"] = ("Das Feld wird überhaupt nur befüllt, wenn der Vorgangstyp leer oder "
                 "\"Makler-Vorgang\" ist UND die Klassifikation \"BÜ-Vorgang\" lautet. "
                 "Mehrere Mängel werden kommagetrennt ausgegeben, z. B. "
                 "\"BÜ-Wunsch fehlt, MV unvollständig\".")
    ws2["A2"].font = Font(name=FONT, sz=10, color="FF4B5563")

    spalten2 = ["Ausprägung", "Wann sie gesetzt wird", "Status"]
    breiten2 = [34, 78, 9]
    kopfzeile(ws2, spalten2, breiten2, zeile=4)
    zeile = 5
    for wert, wann, mark in PRUEFUNG:
        werte = [rein(wert), rein(wann), STATUS_TEXT[mark]]
        schreibe(ws2, zeile, werte, fill=STATUS_FARBE[mark], bold_erste=True)
        hoehe(ws2, zeile, werte, breiten2)
        zeile += 1
    ws2.sheet_view.showGridLines = False

    # ---- Blatt 3: Offene Punkte -----------------------------------------
    ws3 = wb.create_sheet("Offene Punkte")
    ws3["A1"] = "Offene Punkte"
    ws3["A1"].font = Font(name=FONT, sz=14, b=True, color=BLAU)
    ws3["A2"] = "Was noch fachlich zu klären ist, bevor Prompt und Deckblatt nachgezogen werden."
    ws3["A2"].font = Font(name=FONT, sz=10, color="FF4B5563")

    spalten3 = ["Nr.", "Thema", "Frage / Hintergrund"]
    breiten3 = [6, 34, 86]
    kopfzeile(ws3, spalten3, breiten3, zeile=4)
    zeile = 5
    for i, (thema, frage) in enumerate(OFFEN, 1):
        werte = [i, rein(thema), rein(frage)]
        schreibe(ws3, zeile, werte, fill=GELB, bold_erste=False)
        ws3.cell(row=zeile, column=2).font = Font(name=FONT, sz=10, b=True)
        hoehe(ws3, zeile, [rein(thema), rein(frage)], breiten3[1:])
        zeile += 1

    zeile += 1
    ws3.cell(row=zeile, column=2, value="Bekannte Besonderheiten bei Agentur-Nr / "
                                        "Personalnummer").font = Font(name=FONT, sz=11, b=True,
                                                                      color=BLAU)
    zeile += 1
    for i, b in enumerate(BESONDERHEITEN, 1):
        werte = [i, "", rein(b)]
        schreibe(ws3, zeile, werte, fill=GRAU)
        hoehe(ws3, zeile, [rein(b)], [breiten3[2]])
        zeile += 1
    ws3.sheet_view.showGridLines = False

    wb.calculation.fullCalcOnLoad = True
    wb.save(ziel)
    return ziel


if __name__ == "__main__":
    ziel = sys.argv[1] if len(sys.argv) > 1 else "Datenfelder-KI-Deckblatt.xlsx"
    baue_xlsx(ziel)
    print("OK - %s geschrieben (%d Felder)"
          % (ziel, sum(len(f) for _, f in ABSCHNITTE)))
