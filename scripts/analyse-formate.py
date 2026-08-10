#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Format-Analyse der extrahierten Nummern in analyse.xlsx.

Liest das Sheet "Analyse" und schreibt zwei neue Sheets:
  * "Format-Analyse"  - Auswertung: welche Schreibweisen/Formate wurden je Feld
                        extrahiert (Muster, Anzahl, Anteil, Beispiele)
  * "Format-Details"  - Hilfsblatt: pro Vorgangszeile die erkannte Maske +
                        Kategorie. Die Zählungen im Analyse-Blatt sind
                        COUNTIF-Formeln auf dieses Blatt (nachrechenbar).

Aufruf:  python3 scripts/analyse-formate.py [datei.xlsx]
"""

import collections
import math
import re
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATEI = sys.argv[1] if len(sys.argv) > 1 else "analyse.xlsx"
QUELL_SHEET = "Analyse"
SHEET_MAIN = "Format-Analyse"
SHEET_DET = "Format-Details"
KOPFZEILEN = 2          # Zeile 1 = Header, Zeile 2 = Beschreibung
FONT = "Aptos"

# Spalten (0-basiert) im Quellsheet
C_NR, C_VNR, C_POOL, C_AG, C_VM, C_KDNR = 0, 4, 8, 13, 14, 15
C_KLASS = 21                      # Spalte 'Klassifikation'
BUE_WERT = "BUe-Vorgang"          # nur diese Vorgaenge werden ausgewertet

BLAU = "FF1E40AF"
LILA = "FF4C1D95"
GRAU = "FFF3F4F6"
GELB = "FFFEF3C7"
ROT = "FFFEE2E2"
GRUEN = "FFDCFCE7"
WEISS = "FFFFFFFF"

THIN = Side(style="thin", color="FFD1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


# --------------------------------------------------------------------------
# Muster-Erkennung
# --------------------------------------------------------------------------
def maske(s):
    """'60008-1364' -> '9x5-9x4'  (9=Ziffer, A=Großbuchstabe, a=Kleinbuchstabe)"""
    zeichen = []
    for ch in s:
        if ch.isdigit():
            zeichen.append("9")
        elif ch.isalpha():
            zeichen.append("A" if ch.isupper() else "a")
        else:
            zeichen.append(ch)
    out, i = "", 0
    while i < len(zeichen):
        j = i
        while j < len(zeichen) and zeichen[j] == zeichen[i]:
            j += 1
        n = j - i
        out += zeichen[i] if n == 1 else "%sx%d" % (zeichen[i], n)
        i = j
    return out


_TRENNER_NAME = {
    "-": "Bindestrich",
    "/": "Schrägstrich",
    ".": "Punkt",
    ",": "Komma",
    " ": "Leerzeichen",
    "_": "Unterstrich",
}


def _token_text(art, n):
    if art == "9":
        return "%d Ziffer%s" % (n, "n" if n > 1 else "")
    if art == "A":
        return "%d Großbuchstabe%s" % (n, "n" if n > 1 else "")
    if art == "a":
        return "%d Kleinbuchstabe%s" % (n, "n" if n > 1 else "")
    if art == " ":
        return "Leerzeichen" if n == 1 else "%d Leerzeichen" % n
    name = _TRENNER_NAME.get(art, "'%s'" % art)
    return name if n == 1 else "%dx %s" % (n, name)


def crit(s):
    """COUNTIF-Kriterium entschärfen: *, ? und ~ sind dort Platzhalter."""
    return s.replace("~", "~~").replace("*", "~*").replace("?", "~?").replace('"', '""')


def beschreibung_von_rohwert(v):
    zeichen = []
    for ch in v:
        if ch.isdigit():
            zeichen.append("9")
        elif ch.isupper():
            zeichen.append("A")
        elif ch.islower():
            zeichen.append("a")
        else:
            zeichen.append(ch)
    teile, i = [], 0
    while i < len(zeichen):
        j = i
        while j < len(zeichen) and zeichen[j] == zeichen[i]:
            j += 1
        teile.append(_token_text(zeichen[i], j - i))
        i = j
    return " + ".join(teile)


def ziffern(s):
    return re.sub(r"\D", "", s)


def trennzeichen(s):
    tz = sorted({ch for ch in s if not ch.isalnum()})
    return "".join("[Leer]" if c == " " else c for c in tz)


def ist_agenturnummer(s):
    """ERGO-Regel aus ergo-vorgang-analyse.hta: 6000... oder 890... = Agentur."""
    z = ziffern(s)
    return len(z) >= 4 and (z.startswith("6000") or z.startswith("890"))


def hat_text(s):
    """Buchstabenrest, der wie Fließtext aussieht (Label nicht abgetrennt)."""
    return bool(re.search(r"[A-Za-zÄÖÜäöü]{3,}", s))


def zahl(n):
    """1234 -> '1.234' (deutsche Tausendertrennung)."""
    return "{:,}".format(int(n)).replace(",", ".")


def pz(x, nk=1):
    """91.4 -> '91,4'"""
    return ("%.*f" % (nk, x)).replace(".", ",")


def normkey(s):
    z = ziffern(s)
    return z.lstrip("0") or z


# --------------------------------------------------------------------------
# Daten lesen
# --------------------------------------------------------------------------
wb = openpyxl.load_workbook(DATEI)
ws = wb[QUELL_SHEET]
daten = []
for idx, row in enumerate(ws.iter_rows(min_row=KOPFZEILEN + 1, values_only=True), KOPFZEILEN + 1):
    hole = lambda c: ("" if row[c] is None else str(row[c]).strip())
    daten.append({
        "zeile": idx,
        "lfd": hole(C_NR),
        "vnr": hole(C_VNR),
        "pool": hole(C_POOL),
        "ag": hole(C_AG),
        "vm": hole(C_VM),
        "kdnr": hole(C_KDNR),
        "klass": hole(C_KLASS),
        "bue": hole(C_KLASS) == BUE_WERT,
    })

N = len(daten)                       # alle Zeilen im Quellsheet
BUE = [d for d in daten if d["bue"]]  # nur BUe-Vorgaenge
N_BUE = len(BUE)


def werte(feld):
    """Werte des Feldes - ausschliesslich aus BUe-Vorgaengen."""
    return [d[feld] for d in BUE if d[feld]]


# --------------------------------------------------------------------------
# Hilfsblatt "Format-Details"
# --------------------------------------------------------------------------
for name in (SHEET_MAIN, SHEET_DET):
    if name in wb.sheetnames:
        del wb[name]

det = wb.create_sheet(SHEET_DET)
DET_HEAD = [
    "Quellzeile", "Lfd_Nr", "Klassifikation", "BUe_Vorgang",
    "Agentur_Nr", "Agentur_Muster", "Agentur_Beschreibung", "Agentur_Ziffern", "Agentur_Trenner",
    "Vermittler_Nr", "Vermittler_Muster", "Vermittler_Beschreibung", "Vermittler_Ziffern", "Vermittler_Trenner",
    "Versicherungsnummer", "VNR_Muster", "VNR_Familie",
    "Kundennummer", "KDNR_Muster",
    "Maklerpool",
]
det.append(DET_HEAD)


def vnr_familie(v):
    """Familie einer Versicherungsnummer: Buchstaben-Präfix + Ziffernanzahl."""
    if not v:
        return ""
    if "," in v or ";" in v:
        return "Mehrere Nummern in einer Zelle"
    m = re.match(r"^\s*([A-Za-z]{1,3})", v)
    praefix = m.group(1).upper() if m else "(ohne Buchstaben)"
    return "%s + %d Ziffern" % (praefix, len(ziffern(v)))


for d in daten:
    det.append([
        d["zeile"], d["lfd"], d["klass"], "ja" if d["bue"] else "nein",
        d["ag"], maske(d["ag"]) if d["ag"] else "", beschreibung_von_rohwert(d["ag"]) if d["ag"] else "",
        len(ziffern(d["ag"])) if d["ag"] else "", trennzeichen(d["ag"]) if d["ag"] else "",
        d["vm"], maske(d["vm"]) if d["vm"] else "", beschreibung_von_rohwert(d["vm"]) if d["vm"] else "",
        len(ziffern(d["vm"])) if d["vm"] else "", trennzeichen(d["vm"]) if d["vm"] else "",
        d["vnr"], maske(d["vnr"]) if d["vnr"] else "", vnr_familie(d["vnr"]),
        d["kdnr"], maske(d["kdnr"]) if d["kdnr"] else "",
        d["pool"],
    ])

DET_LAST = 1 + N
for c in range(1, len(DET_HEAD) + 1):
    kopfzelle = det.cell(row=1, column=c)
    kopfzelle.font = Font(name=FONT, sz=10, b=True, color=WEISS)
    kopfzelle.fill = PatternFill("solid", fgColor=LILA)
    kopfzelle.alignment = Alignment(vertical="center", wrap_text=True)
    det.column_dimensions[get_column_letter(c)].width = max(11, min(30, len(DET_HEAD[c - 1]) + 3))
det.freeze_panes = "E2"
det.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(DET_HEAD)), DET_LAST)
for r in range(2, DET_LAST + 1):
    for c in range(1, len(DET_HEAD) + 1):
        det.cell(row=r, column=c).font = Font(name=FONT, sz=10)

# Spaltenbuchstaben im Detailblatt (für die COUNTIF-Formeln)
SP = {h: get_column_letter(i) for i, h in enumerate(DET_HEAD, 1)}


def rng(spalte):
    return "'%s'!$%s$2:$%s$%d" % (SHEET_DET, SP[spalte], SP[spalte], DET_LAST)


def cif(paare):
    """COUNTIFS ueber die Detailzeilen - immer eingeschraenkt auf BUe-Vorgaenge."""
    teile = ["%s,%s" % (rng(sp), kriterium) for sp, kriterium in paare]
    teile.append('%s,"ja"' % rng("BUe_Vorgang"))
    return "=COUNTIFS(%s)" % ",".join(teile)


# --------------------------------------------------------------------------
# Hauptblatt "Format-Analyse"
# --------------------------------------------------------------------------
ms = wb.create_sheet(SHEET_MAIN, 1)
BREITEN = [26, 44, 10, 10, 10, 14, 52]
for i, b in enumerate(BREITEN, 1):
    ms.column_dimensions[get_column_letter(i)].width = b

zeile = 1


def setz(r, c, wert, *, bold=False, sz=11, farbe=None, fill=None, fmt=None,
         wrap=False, rahmen=False, italic=False, align=None):
    z = ms.cell(row=r, column=c, value=wert)
    z.font = Font(name=FONT, sz=sz, b=bold, i=italic, color=farbe or "FF111827")
    if fill:
        z.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        z.number_format = fmt
    if wrap or align:
        z.alignment = Alignment(wrap_text=wrap, vertical="top", horizontal=align)
    if rahmen:
        z.border = BORDER
    return z


def titel(text, unterzeile=""):
    global zeile
    zeile += 1
    for c in range(1, 8):
        setz(zeile, c, text if c == 1 else None, bold=True, sz=12, farbe=WEISS, fill=BLAU)
    zeile += 1
    if unterzeile:
        setz(zeile, 1, unterzeile, sz=9, farbe="FF4B5563", fill=GRAU)
        for c in range(2, 8):
            setz(zeile, c, None, fill=GRAU)
        zeile += 1
    return zeile


def tabellenkopf(spalten):
    global zeile
    for i, s in enumerate(spalten, 1):
        setz(zeile, i, s, bold=True, sz=10, farbe=WEISS, fill=LILA, wrap=True, rahmen=True)
    zeile += 1


# ---- Kopf ---------------------------------------------------------------
setz(zeile, 1, "Format-Analyse der extrahierten Nummern", bold=True, sz=16, farbe=BLAU)
zeile += 1
setz(zeile, 1, "Welche Schreibweisen wurden bei Agentur-Nr, Vermittler-Nr, Versicherungsnummer "
               "und Kundennummer tatsächlich extrahiert?", sz=10, farbe="FF4B5563")
zeile += 1
setz(zeile, 1, "Basis: nur BÜ-Vorgänge. Von den %s Zeilen im Sheet '%s' sind %s als "
               "Klassifikation '%s' erfasst - nur diese werden hier ausgewertet, die "
               "übrigen %s (Antrag/Änderung, Anfrage, Kein-Makler-Vorgang, ohne "
               "Klassifikation) bleiben außen vor."
     % (zahl(N), QUELL_SHEET, zahl(N_BUE), BUE_WERT, zahl(N - N_BUE)), sz=10, farbe="FF4B5563")
zeile += 1
setz(zeile, 1, "Alle Anzahlen sind COUNTIFS-Formeln auf das Hilfsblatt '%s' und immer auf "
               "BUe_Vorgang='ja' eingeschränkt." % SHEET_DET, sz=10, farbe="FF4B5563")
zeile += 2

# ---- Kernaussagen -------------------------------------------------------
_ag, _vm, _vnr = werte("ag"), werte("vm"), werte("vnr")


def _mehrfachschreibweisen(vals):
    """(betroffene Nummern, betroffene Vorgänge) fuer Nummern mit >1 Schreibweise."""
    g = collections.defaultdict(collections.Counter)
    for v in vals:
        g[normkey(v)][v] += 1
    multi = {k: c for k, c in g.items() if len(c) > 1}
    return len(multi), sum(sum(c.values()) for c in multi.values())


_ag_multi_n, _ag_multi_v = _mehrfachschreibweisen(_ag)
_vm_multi_n, _vm_multi_v = _mehrfachschreibweisen(_vm)
_ag_top2 = sum(n for _, n in collections.Counter(maske(v) for v in _ag).most_common(2))
_vm_top1 = collections.Counter(maske(v) for v in _vm).most_common(1)[0][1]
_ag_nur_ziffern = sum(1 for v in _ag if v.isdigit())
_vm_nur_ziffern = sum(1 for v in _vm if v.isdigit())
_vm_nullen = sum(1 for v in _vm if ziffern(v).startswith("0"))
_defekt = sum(1 for v in _ag + _vm
              if hat_text(v) or v.startswith("{") or len(ziffern(v)) < 5 or len(ziffern(v)) > 9)

KERN = [
    ("Agentur_Nr ist fast einheitlich",
     "%s von %s Werten (%s %%) stehen in nur 2 kompakten Formaten: 7 Ziffern mit 890-Präfix "
     "und 9 Ziffern mit 6000-Präfix. Insgesamt nur %d verschiedene Muster."
     % (zahl(_ag_top2), zahl(len(_ag)), pz(100.0 * _ag_top2 / len(_ag)),
        len({maske(v) for v in _ag}))),
    ("Vermittler_Nr ist das Auffangfeld",
     "%d verschiedene Muster bei nur %s Werten. Häufigster Fall: 6 Ziffern kompakt (%s Werte, "
     "%s %%) - der Rest verteilt sich auf Bindestrich-, Schrägstrich-, Leerzeichen- und "
     "Buchstaben-Varianten."
     % (len({maske(v) for v in _vm}), zahl(len(_vm)), zahl(_vm_top1), pz(100.0 * _vm_top1 / len(_vm)))),
    ("Trennzeichen sind der Hauptunterschied, nicht die Nummer",
     "%s %% der Agentur- und %s %% der Vermittler-Werte sind reine Ziffernketten. Die "
     "übrigen enthalten dieselbe Nummer nur mit Bindestrich, Schrägstrich, Leerzeichen oder "
     "Buchstaben-Präfix."
     % (pz(100.0 * _ag_nur_ziffern / len(_ag)), pz(100.0 * _vm_nur_ziffern / len(_vm)))),
    ("Dieselbe Nummer erscheint mehrfach verschieden geschrieben",
     "%d Agenturnummern (%s Vorgänge) und %d Vermittlernummern (%s Vorgänge) kommen in "
     "mindestens zwei Schreibweisen vor - z. B. 60008-1364 neben 600081364. Ohne "
     "Normalisierung zersplittert das Cluster und Makler-Auswertungen. Details in Block 5."
     % (_ag_multi_n, zahl(_ag_multi_v), _vm_multi_n, zahl(_vm_multi_v))),
    ("Führende Nullen nur im Vermittler-Feld",
     "%s Vermittler-Werte beginnen mit einer 0 (008923555 vs. 8923555). Bei Agentur_Nr kommt "
     "das nicht vor." % zahl(_vm_nullen)),
    ("Echte Extraktionsfehler sind selten",
     "%d von %s Nummernwerten (%s %%) sind formal defekt: Textrest, Platzhalter, zu kurz "
     "oder zu lang. Die Extraktion selbst ist also stabil - der Aufwand liegt in der "
     "Vereinheitlichung, nicht in der Erkennung."
     % (_defekt, zahl(len(_ag) + len(_vm)), pz(100.0 * _defekt / (len(_ag) + len(_vm)), 2))),
    ("Versicherungsnummer ist das heterogenste Feld",
     "%d verschiedene Rohmuster bei %s Werten, zusammengefasst zu %d Formatfamilien "
     "(Buchstaben-Präfix + Ziffernanzahl). Dominant: KV/SV/LF/KR/LV mit 8 oder 9 Ziffern. "
     "Siehe Block 7."
     % (len({maske(v) for v in _vnr}), zahl(len(_vnr)), len({vnr_familie(v) for v in _vnr}))),
]

titel("Kernaussagen", "Das Wichtigste aus den Blöcken 0 bis 8 auf einen Blick.")
for i, (kopf, text) in enumerate(KERN, 1):
    setz(zeile, 1, "%d. %s" % (i, kopf), bold=True, sz=10, wrap=True, rahmen=True, fill=GRAU)
    setz(zeile, 2, text, sz=10, wrap=True, rahmen=True)
    for c in range(3, 8):
        setz(zeile, c, None, rahmen=True)
    ms.merge_cells(start_row=zeile, start_column=2, end_row=zeile, end_column=7)
    ms.row_dimensions[zeile].height = 14 * max(2, math.ceil(len(text) / 130.0))
    zeile += 1
zeile += 1


# ---- Verwendeter Prompt + Nachverarbeitung ------------------------------
# Wortlaut aus ergo-vorgang-analyse.hta (buildPrompt / wsBuildBatchPrompt +
# deriveMaklernummern). Der Prompt selbst ist ASCII-transliteriert (ue/ae/oe),
# weil er genau so an das Modell geht - deshalb hier unveraendert uebernommen.
PROMPT_AG = (
    "[QUELLE: BUe-Wunsch/Anschreiben] 3a1) agentur_nummer - NUR die Agentur-Nr als "
    "REINE NUMMER/Kennung OHNE Text (z.B. '8903235'). Leer wenn nicht genannt."
)
PROMPT_VM = (
    "[QUELLE: BUe-Wunsch/Anschreiben] 3a2) vermittler_nummer - NUR die Vermittler-/"
    "Personalnummer als REINE NUMMER/Kennung OHNE Text (z.B. '811774'). Sie ist i.d.R. "
    "6-stellig (Personalnummer). Leer wenn nicht genannt."
)
PROMPT_GEMEINSAM = (
    "WICHTIG: agentur_nummer und vermittler_nummer enthalten AUSSCHLIESSLICH die "
    "Ziffern/Kennung - KEINE Woerter wie 'Agentur'/'Vermittler' und KEINE Vermischung "
    "beider Nummern in einem Feld."
)
PROMPT_PRIO = (
    "[QUELLE: BUe-Wunsch/Anschreiben] 3a) makler_nummer - Identifikator BEIM VERSICHERER. "
    "STRENGE Priorisierung:\n"
    "    1. 'Agentur-Nr.' / 'Vermittler-Nr.' im Briefkopf-Block (das ist die Wunsch-Nummer). "
    "HINWEIS: die Agentur-/Vermittlernummer beginnt oft mit '6000' oder '890'.\n"
    "    2. 'BD-Nummer' / 'Bestandsnummer' / 'Beraterzeichen' / 'PNR' / 'Personalnummer'.\n"
    "    3. NUR Fallback: Pool-IDs wie 'MAKID','DEMMAK','BCA-ID' (sind pool-intern, NICHT "
    "beim Versicherer)."
)
PROMPT_SCHEMA = (
    '"agentur_nummer":"[QUELLE: BUe-Wunsch/Anschreiben] NUR der reine Agentur-Nummern-Wert '
    'OHNE das Wort Agentur. WICHTIG: jede Nummer, die mit 6000 oder 890 beginnt, ist eine '
    'AGENTURNUMMER und gehoert IMMER hierher (NICHT nach vermittler_nummer), z.B. 8903235, '
    '600042497; leer wenn keine"\n'
    '"vermittler_nummer":"[QUELLE: BUe-Wunsch/Anschreiben] NUR der reine Vermittler-/'
    'BD-Nummern-Wert OHNE das Wort Vermittler. NUR Nummern, die NICHT mit 6000/890 beginnen '
    '(z.B. 811774 oder 898-0003); eine 6000.../890...-Nummer NIEMALS hier ablegen; leer '
    'wenn keine"'
)

_textrest = [v for v in _ag + _vm if hat_text(v)]
_platzhalter = [v for v in _ag + _vm if v.startswith("{")]

PROMPT_BLOCK = [
    ("Prompt: Feld agentur_nummer", PROMPT_AG, GRAU),
    ("Prompt: Feld vermittler_nummer", PROMPT_VM, GRAU),
    ("Prompt: gemeinsame Regel", PROMPT_GEMEINSAM, GRAU),
    ("Prompt: Priorisierung der Quelle", PROMPT_PRIO, GRAU),
    ("Prompt: JSON-Schema im Batch-Lauf", PROMPT_SCHEMA, GRAU),
    ("Nachverarbeitung im Tool",
     "deriveMaklernummern() greift nach der KI-Antwort noch einmal ein: (1) Platzhalter wie "
     "'leer', '-', 'k.A.', 'n/a', 'keine' werden geleert. (2) Fehlt ein Feld, wird die "
     "kombinierte makler_nummer per Regex zerlegt (Muster: Label + optionaler Buchstabe + "
     "Ziffern, Bindestrich und Schrägstrich erlaubt). (3) Doppelte Nummern innerhalb "
     "desselben Vorgangs werden über den Ziffernvergleich entfernt, 008923555 gilt als "
     "identisch mit 8923555. (4) Die 6000/890-Regel entscheidet abschließend, in welche "
     "Spalte ein Wert kommt - unabhängig davon, in welchem KI-Feld er stand.", None),
    ("Was dadurch schon vereinheitlicht ist",
     "Die Spaltenzuordnung: alle %s Agentur-Werte erfüllen die 6000/890-Regel, keine "
     "Vermischung mehr. Außerdem sind Label-Texte in aller Regel entfernt und "
     "Standard-Platzhalter geleert."
     % zahl(sum(1 for v in _ag if ist_agenturnummer(v))), GRUEN),
    ("Was der Prompt NICHT regelt",
     "Weder Prompt noch Nachverarbeitung schreiben eine Schreibweise vor. Bindestrich, "
     "Schrägstrich, Leerzeichen, führende Nullen und Buchstaben-Präfixe werden 1:1 aus dem "
     "Dokument übernommen - der Regex im Parser lässt Buchstaben, '-' und '/' sogar "
     "ausdrücklich zu. Auch eine Sollstellenzahl (6-stellig, 7-stellig, 9-stellig) wird "
     "nirgends erzwungen.", GELB),
    ("Wo der Prompt nicht durchgegriffen hat",
     "%d Werte enthalten trotz der Regel noch Wortbestandteile - Label, Verbindungswort "
     "oder Kürzel-Zusatz (%s). %d Wert ist ein unersetzter "
     "Feldname (%s) - dieser Platzhalter steht nicht in der Filterliste des Parsers und "
     "rutscht deshalb durch."
     % (len(_textrest), ", ".join(sorted(set(_textrest))[:4]),
        len(_platzhalter), ", ".join(sorted(set(_platzhalter))) or "keiner"), ROT),
    ("Ansatzpunkt",
     "Die Formate im Rohdokument lassen sich nicht steuern - wohl aber die Ablage. Eine "
     "Normalisierung in deriveMaklernummern() (nur Ziffern behalten, auf die Sollstellenzahl "
     "auffüllen, Buchstaben-Präfix separat führen) würde die in Block 5 gelisteten Dubletten "
     "auflösen, ohne dass der Prompt angefasst werden muss.", None),
]

titel("Verwendeter Prompt für Agentur- und Vermittlernummer",
      "Wortlaut aus ergo-vorgang-analyse.hta. Der Prompt ist bewusst ohne Umlaute "
      "geschrieben (ue/ae/oe), weil er genau so an das Modell geht.")
for kopf, text, farbe in PROMPT_BLOCK:
    setz(zeile, 1, kopf, bold=True, sz=10, wrap=True, rahmen=True, fill=GRAU)
    setz(zeile, 2, text, sz=10, wrap=True, rahmen=True, fill=farbe)
    for c in range(3, 8):
        setz(zeile, c, None, rahmen=True, fill=farbe)
    ms.merge_cells(start_row=zeile, start_column=2, end_row=zeile, end_column=7)
    umbrueche = text.count("\n")
    ms.row_dimensions[zeile].height = 14 * max(2, math.ceil(len(text) / 130.0) + umbrueche)
    zeile += 1
zeile += 1

# ---- Block 0: Befüllung ------------------------------------------------
titel("0 - Befüllung der Nummernfelder",
      "Wie oft hat die Extraktion überhaupt einen Wert geliefert? Bezugsgröße sind die "
      "%s BÜ-Vorgänge." % zahl(N_BUE))
tabellenkopf(["Feld", "Beschreibung", "gefüllt", "Anteil", "leer", "verschiedene Werte", "verschiedene Muster"])

BEFUELLUNG = [
    ("Agentur_Nr", "Agentur-/Vermittlernummer mit Präfix 6000 oder 890 (ERGO-Regel im Tool)",
     "Agentur_Nr", "Agentur_Muster"),
    ("Vermittler_Nr", "Personal-/Vermittlernummer - alles ohne 6000-/890-Präfix",
     "Vermittler_Nr", "Vermittler_Muster"),
    ("Versicherungsnummer", "Vertrags-/Versicherungsnummer des Endkunden",
     "Versicherungsnummer", "VNR_Muster"),
    ("Kundennummer", "Kunden-/Partnernummer (KDNR) in ERGO-Systemen",
     "Kundennummer", "KDNR_Muster"),
]
befuellung_start = zeile
for label, beschr, sp_wert, sp_maske in BEFUELLUNG:
    feld = {"Agentur_Nr": "ag", "Vermittler_Nr": "vm",
            "Versicherungsnummer": "vnr", "Kundennummer": "kdnr"}[label]
    vals = werte(feld)
    setz(zeile, 1, label, bold=True, sz=10, rahmen=True)
    setz(zeile, 2, beschr, sz=10, wrap=True, rahmen=True)
    setz(zeile, 3, cif([(sp_wert, '"<>"')]), sz=10, fmt="#,##0", rahmen=True)
    setz(zeile, 4, "=C%d/%d" % (zeile, N_BUE), sz=10, fmt="0.0%", rahmen=True)
    setz(zeile, 5, "=%d-C%d" % (N_BUE, zeile), sz=10, fmt="#,##0", rahmen=True)
    setz(zeile, 6, len(set(vals)), sz=10, fmt="#,##0", rahmen=True)
    setz(zeile, 7, len({maske(v) for v in vals}), sz=10, fmt="#,##0", rahmen=True)
    zeile += 1

setz(zeile, 1, "Die Zeilen oben überschneiden sich (ein Vorgang kann beide Nummern haben) "
               "und ergeben deshalb zusammen nicht 100 %%. Die Matrix darunter teilt dieselben "
               "%s Vorgänge überschneidungsfrei auf." % zahl(N_BUE),
     sz=9, farbe="FF4B5563")
zeile += 2

# ---- Block 0b: Matrix Agentur x Personalnummer --------------------------
titel("0b - Matrix: Agentur-Nr x Personalnummer",
      "Jeder der %s BÜ-Vorgänge fällt in genau eines der vier inneren Felder - "
      "die Anteile addieren sich zu 100 %%." % zahl(N_BUE))
tabellenkopf(["", "Personalnummer vorhanden", "Personalnummer fehlt", "SUMME",
              "", "", ""])

matrix_start = zeile
# Zeile 1: Agentur vorhanden
setz(zeile, 1, "Agentur-Nr vorhanden", bold=True, sz=10, rahmen=True, fill=GRAU)
setz(zeile, 2, cif([("Agentur_Nr", '"<>"'), ("Vermittler_Nr", '"<>"')]),
     sz=10, fmt="#,##0", rahmen=True, fill=GRUEN)
setz(zeile, 3, cif([("Agentur_Nr", '"<>"'), ("Vermittler_Nr", '""')]),
     sz=10, fmt="#,##0", rahmen=True)
setz(zeile, 4, "=B%d+C%d" % (zeile, zeile), bold=True, sz=10, fmt="#,##0",
     rahmen=True, fill=GRAU)
zeile += 1
# Zeile 2: Agentur fehlt
setz(zeile, 1, "Agentur-Nr fehlt", bold=True, sz=10, rahmen=True, fill=GRAU)
setz(zeile, 2, cif([("Agentur_Nr", '""'), ("Vermittler_Nr", '"<>"')]),
     sz=10, fmt="#,##0", rahmen=True)
setz(zeile, 3, cif([("Agentur_Nr", '""'), ("Vermittler_Nr", '""')]),
     sz=10, fmt="#,##0", rahmen=True, fill=ROT)
setz(zeile, 4, "=B%d+C%d" % (zeile, zeile), bold=True, sz=10, fmt="#,##0",
     rahmen=True, fill=GRAU)
zeile += 1
# Summenzeile
setz(zeile, 1, "SUMME", bold=True, sz=10, rahmen=True, fill=GRAU)
for sp in ("B", "C", "D"):
    setz(zeile, "ABCD".index(sp) + 1,
         "=%s%d+%s%d" % (sp, matrix_start, sp, matrix_start + 1),
         bold=True, sz=10, fmt="#,##0", rahmen=True, fill=GRAU)
zeile += 2

# Dieselbe Matrix in Prozent. Damit die vier Felder exakt 100 % ergeben,
# wird ein Feld als Rest gerechnet (1 minus die drei anderen) - so stimmen
# auch die Rand- und Gesamtsummen ohne Rundungsdifferenz.
tabellenkopf(["Anteil an allen %s BÜ-Vorgängen" % zahl(N_BUE),
              "Personalnummer vorhanden", "Personalnummer fehlt", "SUMME", "", "", ""])
proz_start = zeile
# Drei Felder auf die angezeigte Genauigkeit gerundet (0,1 Prozentpunkt =
# 3 Nachkommastellen), das vierte als Rest -> die vier Felder ergeben in der
# Anzeige exakt 100 %, und auch die Rand- und Gesamtsummen gehen auf.
setz(zeile, 1, "Agentur-Nr vorhanden", bold=True, sz=10, rahmen=True, fill=GRAU)
setz(zeile, 2, "=ROUND(B%d/$D$%d,3)" % (matrix_start, matrix_start + 2), sz=10,
     fmt="0.0%", rahmen=True, fill=GRUEN)
setz(zeile, 3, "=1-B%d-B%d-C%d" % (proz_start, proz_start + 1, proz_start + 1),
     sz=10, fmt="0.0%", rahmen=True)
setz(zeile, 4, "=B%d+C%d" % (zeile, zeile), bold=True, sz=10, fmt="0.0%",
     rahmen=True, fill=GRAU)
zeile += 1
setz(zeile, 1, "Agentur-Nr fehlt", bold=True, sz=10, rahmen=True, fill=GRAU)
setz(zeile, 2, "=ROUND(B%d/$D$%d,3)" % (matrix_start + 1, matrix_start + 2), sz=10,
     fmt="0.0%", rahmen=True)
setz(zeile, 3, "=ROUND(C%d/$D$%d,3)" % (matrix_start + 1, matrix_start + 2), sz=10,
     fmt="0.0%", rahmen=True, fill=ROT)
setz(zeile, 4, "=B%d+C%d" % (zeile, zeile), bold=True, sz=10, fmt="0.0%",
     rahmen=True, fill=GRAU)
zeile += 1
setz(zeile, 1, "SUMME", bold=True, sz=10, rahmen=True, fill=GRAU)
for sp in ("B", "C", "D"):
    setz(zeile, "ABCD".index(sp) + 1,
         "=%s%d+%s%d" % (sp, proz_start, sp, proz_start + 1),
         bold=True, sz=10, fmt="0.0%", rahmen=True, fill=GRAU)
zeile += 1
setz(zeile, 1, "Lesehilfe: grün = beide Nummern vorhanden, rot = gar keine Maklernummer. "
               "Drei Felder sind auf 0,1 Prozentpunkt gerundet, das Feld "
               "'Agentur vorhanden / Personalnummer fehlt' ist der Rest zu 100 % - dadurch "
               "ergeben die vier Felder in der Anzeige exakt 100 %, und die Rand- und "
               "Gesamtsummen gehen ohne Rundungsdifferenz auf.",
     sz=9, farbe="FF4B5563")
zeile += 2


# ---- Format-Tabelle generisch -------------------------------------------
def formattabelle(feld, sp_wert, sp_maske, ueberschrift, hinweis, regelspalte=None):
    """Muster-Häufigkeiten für ein Feld ausgeben."""
    global zeile
    vals = werte(feld)
    cnt = collections.Counter(maske(v) for v in vals)
    beispiele = collections.defaultdict(list)
    for v in vals:
        m = maske(v)
        if len(beispiele[m]) < 3 and v not in beispiele[m]:
            beispiele[m].append(v)

    titel(ueberschrift, hinweis)
    kopf = ["Muster (9=Ziffer, A=Buchstabe)", "Format in Worten", "Anzahl", "Anteil",
            "Ziffern", "Trennzeichen", "Beispiele aus den Daten"]
    if regelspalte:
        kopf[5] = regelspalte
    tabellenkopf(kopf)
    start = zeile
    for m, n in cnt.most_common():
        bsp = beispiele[m]
        zif = len(ziffern(bsp[0]))
        tz = trennzeichen(bsp[0]) or "-"
        if regelspalte:
            treffer = all(ist_agenturnummer(b) for b in bsp)
            tz = "ja" if treffer else "nein"
        auffällig = (n <= 2) or hat_text(bsp[0]) or bsp[0].startswith("{")
        f = GELB if auffällig else None
        setz(zeile, 1, m, sz=10, rahmen=True, fill=f)
        setz(zeile, 2, beschreibung_von_rohwert(bsp[0]), sz=10, wrap=True, rahmen=True, fill=f)
        # Rein numerische Masken (z.B. "9") wuerde COUNTIF als Zahl interpretieren -
        # dort exakter Textvergleich per SUMPRODUCT/EXACT.
        formel = ('=SUMPRODUCT(--EXACT(%s,"%s"),--(%s="ja"))'
                  % (rng(sp_maske), m, rng("BUe_Vorgang"))) if m.isdigit() \
            else cif([(sp_maske, '"%s"' % crit(m))])
        setz(zeile, 3, formel, sz=10, fmt="#,##0", rahmen=True, fill=f)
        setz(zeile, 4, '=C%d/%s' % (zeile, cif([(sp_wert, '"<>"')])[1:]), sz=10,
             fmt="0.0%", rahmen=True, fill=f)
        setz(zeile, 5, zif, sz=10, fmt="0", rahmen=True, fill=f)
        setz(zeile, 6, tz, sz=10, rahmen=True, fill=f)
        setz(zeile, 7, ", ".join(bsp), sz=10, wrap=True, rahmen=True, fill=f)
        zeile += 1
    setz(zeile, 1, "SUMME", bold=True, sz=10, rahmen=True, fill=GRAU)
    setz(zeile, 2, "%d verschiedene Muster" % len(cnt), bold=True, sz=10, rahmen=True, fill=GRAU)
    setz(zeile, 3, "=SUM(C%d:C%d)" % (start, zeile - 1), bold=True, sz=10, fmt="#,##0",
         rahmen=True, fill=GRAU)
    setz(zeile, 4, "=SUM(D%d:D%d)" % (start, zeile - 1), bold=True, sz=10, fmt="0.0%",
         rahmen=True, fill=GRAU)
    for c in (5, 6, 7):
        setz(zeile, c, None, fill=GRAU, rahmen=True)
    zeile += 1
    setz(zeile, 1, "Gelb hinterlegt = Einzelfall (max. 2 Treffer) oder Textrest/Platzhalter im Wert.",
         sz=9, farbe="FF92400E")
    zeile += 2


formattabelle(
    "ag", "Agentur_Nr", "Agentur_Muster",
    "1 - Agentur_Nr: extrahierte Formate",
    "Das Tool sortiert jede Nummer, deren Ziffern mit 6000 oder 890 beginnen, in dieses Feld "
    "(deriveMaklernummern in ergo-vorgang-analyse.hta). Diese Regel greift sauber: %s von %s "
    "Werten (%s %%) erfüllen sie. Die Frage ist also nicht die Zuordnung, sondern die "
    "Schreibweise."
    % (zahl(sum(1 for v in werte("ag") if ist_agenturnummer(v))), zahl(len(werte("ag"))),
       pz(100.0 * sum(1 for v in werte("ag") if ist_agenturnummer(v)) / len(werte("ag")))))

formattabelle(
    "vm", "Vermittler_Nr", "Vermittler_Muster",
    "2 - Vermittler_Nr: extrahierte Formate",
    "Auffangfeld: alles, was nicht mit 6000/890 beginnt. Entsprechend heterogen - "
    "Personalnummern, Beraterzeichen, Pool-IDs, Untervermittler-Kennungen.")


# ---- Block 3: Ziffernlängen -------------------------------------------
titel("3 - Ziffernlängen (Trennzeichen und Buchstaben herausgerechnet)",
      "Zeigt, welche Nummernlängen fachlich vorkommen - unabhängig von der Schreibweise.")
tabellenkopf(["Ziffern gesamt", "Bedeutung / typischer Fall", "Agentur_Nr", "Anteil",
              "Vermittler_Nr", "Anteil", "Beispiele"])

len_ag = collections.Counter(len(ziffern(v)) for v in werte("ag"))
len_vm = collections.Counter(len(ziffern(v)) for v in werte("vm"))
bsp_len = collections.defaultdict(list)
for v in werte("ag") + werte("vm"):
    k = len(ziffern(v))
    if len(bsp_len[k]) < 3 and v not in bsp_len[k]:
        bsp_len[k].append(v)

BEDEUTUNG = {
    5: "verkürzte Vermittlernummer (führende Nullen fehlen)",
    6: "Standard-Personalnummer / Vermittlernummer",
    7: "Agenturnummer im 890-Format (890 + 4 Stellen)",
    8: "Vermittlernummer mit führender Null oder Pool-ID",
    9: "Agenturnummer im 6000-Format (6000 + 5 Stellen) bzw. 9-stellige Pool-ID",
}
for k in sorted(set(len_ag) | set(len_vm)):
    warn = k < 5 or k > 9
    f = ROT if warn else None
    setz(zeile, 1, k, sz=10, fmt="0", rahmen=True, fill=f)
    setz(zeile, 2, BEDEUTUNG.get(k, "unplausibel - Fragment, Dopplung oder Textrest"),
         sz=10, wrap=True, rahmen=True, fill=f)
    setz(zeile, 3, cif([("Agentur_Ziffern", str(k))]), sz=10, fmt="#,##0",
         rahmen=True, fill=f)
    setz(zeile, 4, '=IFERROR(C%d/%s,0)' % (zeile, cif([("Agentur_Nr", '"<>"')])[1:]),
         sz=10, fmt="0.0%", rahmen=True, fill=f)
    setz(zeile, 5, cif([("Vermittler_Ziffern", str(k))]), sz=10, fmt="#,##0",
         rahmen=True, fill=f)
    setz(zeile, 6, '=IFERROR(E%d/%s,0)' % (zeile, cif([("Vermittler_Nr", '"<>"')])[1:]),
         sz=10, fmt="0.0%", rahmen=True, fill=f)
    setz(zeile, 7, ", ".join(bsp_len[k]), sz=10, wrap=True, rahmen=True, fill=f)
    zeile += 1
setz(zeile, 1, "Rot = Länge außerhalb 5-9 Stellen, faktisch immer ein Extraktionsfehler.",
     sz=9, farbe="FF991B1B")
zeile += 2


# ---- Block 4: Trennzeichen ----------------------------------------------
titel("4 - Trennzeichen-Varianten derselben Nummer",
      "Kompakt (600081364) vs. getrennt (60008-1364 / 60008/0396) - dieselbe Nummer, "
      "verschiedene Schreibweisen.")
tabellenkopf(["Schreibweise", "Erkennungsmerkmal", "Agentur_Nr", "Anteil",
              "Vermittler_Nr", "Anteil", "Beispiele"])

VARIANTEN = [
    ("kompakt (nur Ziffern)", "keine Trennzeichen, keine Buchstaben", ""),
    ("mit Bindestrich", "z.B. 60008-1364, 890-3235, 333-6272", "-"),
    ("mit Schrägstrich", "z.B. 60002/0753, 893/4055", "/"),
    ("mit Leerzeichen", "z.B. 'A 600046012', '00 111 40 16'", " "),
    ("mit Buchstaben-Präfix", "z.B. A600040124, V 85357, P42356", "ALPHA"),
    ("mit Textrest im Wert", "Label nicht abgetrennt, z.B. 'Vermittler 891300'", "TEXT"),
]
for label, merkmal, key in VARIANTEN:
    def zaehle(feld):
        n = 0
        for v in werte(feld):
            if key == "":
                if v.isdigit():
                    n += 1
            elif key == "ALPHA":
                if re.search(r"[A-Za-z]", v) and not hat_text(v):
                    n += 1
            elif key == "TEXT":
                if hat_text(v):
                    n += 1
            elif key in v:
                n += 1
        return n

    bsp = []
    for v in werte("ag") + werte("vm"):
        ok = (v.isdigit() if key == "" else
              (re.search(r"[A-Za-z]", v) and not hat_text(v)) if key == "ALPHA" else
              hat_text(v) if key == "TEXT" else key in v)
        if ok and len(bsp) < 3 and v not in bsp:
            bsp.append(v)
    f = ROT if key == "TEXT" else None
    setz(zeile, 1, label, bold=True, sz=10, rahmen=True, fill=f)
    setz(zeile, 2, merkmal, sz=10, wrap=True, rahmen=True, fill=f)
    setz(zeile, 3, zaehle("ag"), sz=10, fmt="#,##0", rahmen=True, fill=f)
    setz(zeile, 4, '=IFERROR(C%d/%s,0)' % (zeile, cif([("Agentur_Nr", '"<>"')])[1:]),
         sz=10, fmt="0.0%", rahmen=True, fill=f)
    setz(zeile, 5, zaehle("vm"), sz=10, fmt="#,##0", rahmen=True, fill=f)
    setz(zeile, 6, '=IFERROR(E%d/%s,0)' % (zeile, cif([("Vermittler_Nr", '"<>"')])[1:]),
         sz=10, fmt="0.0%", rahmen=True, fill=f)
    setz(zeile, 7, ", ".join(bsp), sz=10, wrap=True, rahmen=True, fill=f)
    zeile += 1
setz(zeile, 1, "Anzahlen in diesem Block sind Werte (Klassifikation per Skript), keine Formeln - "
               "die Zuordnung lässt sich in '%s' Spalte Trenner nachvollziehen." % SHEET_DET,
     sz=9, farbe="FF4B5563")
zeile += 2


# ---- Block 5: gleiche Nummer, mehrere Schreibweisen ---------------------
titel("5 - Dieselbe Nummer in mehreren Schreibweisen",
      "Nach Normalisierung (nur Ziffern, führende Nullen weg) identisch, im Rohwert aber "
      "verschieden. Das zersplittert Cluster und Makler-Auswertungen.")
tabellenkopf(["Normalisierte Nummer", "Feld", "gefundene Schreibweisen", "Varianten",
              "Vorgänge gesamt", "", ""])

for feld, sp_wert, label in (("ag", "Agentur_Nr", "Agentur_Nr"), ("vm", "Vermittler_Nr", "Vermittler_Nr")):
    gruppen = collections.defaultdict(collections.Counter)
    for v in werte(feld):
        gruppen[normkey(v)][v] += 1
    multi = {k: c for k, c in gruppen.items() if len(c) > 1}
    for k, c in sorted(multi.items(), key=lambda x: (-sum(x[1].values()), x[0])):
        setz(zeile, 1, k, sz=10, rahmen=True, fill=GELB)
        setz(zeile, 2, label, sz=10, rahmen=True, fill=GELB)
        setz(zeile, 3, "  |  ".join("%s (%dx)" % (v, n) for v, n in c.most_common()),
             sz=10, wrap=True, rahmen=True, fill=GELB)
        setz(zeile, 4, len(c), sz=10, fmt="0", rahmen=True, fill=GELB)
        setz(zeile, 5, sum(c.values()), sz=10, fmt="#,##0", rahmen=True, fill=GELB)
        zeile += 1
zeile += 1


# ---- Block 6: Auffälligkeiten -----------------------------------------
titel("6 - Auffällige Extraktionen (Prüfliste)",
      "Werte, die formal nicht zu einer Makler-/Agenturnummer passen - Kandidaten für "
      "Nachschärfung von Prompt oder Parser.")
tabellenkopf(["Kategorie", "Warum auffällig", "Agentur_Nr", "Vermittler_Nr", "",
              "", "Betroffene Werte"])

def sammle(pruef):
    res = {}
    for feld in ("ag", "vm"):
        treffer = [v for v in werte(feld) if pruef(v)]
        res[feld] = treffer
    return res

PRUEFUNGEN = [
    ("Textrest im Wert", "Label wie 'Vermittler'/'Agentur' oder Firmenzusatz nicht abgetrennt",
     lambda v: hat_text(v)),
    ("Platzhalter", "Modell hat den Feldnamen statt eines Wertes geliefert",
     lambda v: v.startswith("{") or v.lower() in ("k.a.", "leer", "-", "n/a")),
    ("zu kurz (<5 Ziffern)", "kann keine vollständige Makler-/Agenturnummer sein",
     lambda v: len(ziffern(v)) < 5),
    ("zu lang (>9 Ziffern)", "meist zwei zusammengezogene Nummern oder Nummer + Zusatz",
     lambda v: len(ziffern(v)) > 9),
    ("mehrere Nummern in einer Zelle", "Komma/'und' im Wert - Feld ist einwertig gedacht",
     lambda v: ("," in v) or (" und " in v.lower())),
    ("führende Nullen", "008923555 vs 8923555 - gleiche Nummer, unterschiedlicher Schlüssel",
     lambda v: ziffern(v).startswith("0")),
]
for label, warum, pruef in PRUEFUNGEN:
    res = sammle(pruef)
    alle = res["ag"] + res["vm"]
    uniq = sorted(set(alle))
    f = ROT if alle else GRUEN
    setz(zeile, 1, label, bold=True, sz=10, rahmen=True, fill=f)
    setz(zeile, 2, warum, sz=10, wrap=True, rahmen=True, fill=f)
    setz(zeile, 3, len(res["ag"]), sz=10, fmt="#,##0", rahmen=True, fill=f)
    setz(zeile, 4, len(res["vm"]), sz=10, fmt="#,##0", rahmen=True, fill=f)
    setz(zeile, 5, None, rahmen=True, fill=f)
    setz(zeile, 6, None, rahmen=True, fill=f)
    setz(zeile, 7, ", ".join(uniq[:12]) + (" ..." if len(uniq) > 12 else ""),
         sz=10, wrap=True, rahmen=True, fill=f)
    zeile += 1
zeile += 1


# ---- Block 7: Versicherungsnummer --------------------------------------
titel("7 - Versicherungsnummer: Formatfamilien",
      "Zusammengefasst nach Buchstaben-Präfix + Ziffernanzahl (Trennzeichen ignoriert). "
      "Die Rohmuster stehen je Zeile im Hilfsblatt '%s'." % SHEET_DET)
tabellenkopf(["Familie", "Format in Worten", "Anzahl", "Anteil", "Muster-Varianten",
              "", "Beispiele"])

fam = collections.Counter(vnr_familie(v) for v in werte("vnr"))
fam_bsp = collections.defaultdict(list)
fam_masken = collections.defaultdict(set)
for v in werte("vnr"):
    k = vnr_familie(v)
    fam_masken[k].add(maske(v))
    if len(fam_bsp[k]) < 3 and v not in fam_bsp[k]:
        fam_bsp[k].append(v)

vnr_start = zeile
for k, n in fam.most_common(30):
    setz(zeile, 1, k, sz=10, rahmen=True)
    setz(zeile, 2, beschreibung_von_rohwert(fam_bsp[k][0]), sz=10, wrap=True, rahmen=True)
    setz(zeile, 3, cif([("VNR_Familie", '"%s"' % crit(k))]), sz=10, fmt="#,##0",
         rahmen=True)
    setz(zeile, 4, '=C%d/%s' % (zeile, cif([("Versicherungsnummer", '"<>"')])[1:]),
         sz=10, fmt="0.0%", rahmen=True)
    setz(zeile, 5, len(fam_masken[k]), sz=10, fmt="0", rahmen=True)
    setz(zeile, 6, None, rahmen=True)
    setz(zeile, 7, ", ".join(fam_bsp[k]), sz=10, wrap=True, rahmen=True)
    zeile += 1
setz(zeile, 1, "Top 30 von %d Familien / %d Rohmustern. 'Muster-Varianten' = wie viele "
               "unterschiedliche Schreibweisen (mit/ohne Leerzeichen, Punkt, Bindestrich) "
               "in dieser Familie vorkommen." % (len(fam), len({maske(v) for v in werte("vnr")})),
     sz=9, farbe="FF4B5563")
zeile += 2


# ---- Block 8: Format je Maklerpool -------------------------------------
titel("8 - Welcher Maklerpool liefert welches Nummernformat?",
      "Top-Pools nach Vorgangsmenge mit ihrem dominierenden Agentur-/Vermittler-Muster.")
tabellenkopf(["Maklerpool", "Vorgänge", "häufigstes Agentur-Muster", "Beispiel",
              "häufigstes Vermittler-Muster", "Beispiel", "Muster-Vielfalt (Ag / Vm)"])

pool_cnt = collections.Counter(d["pool"] for d in BUE if d["pool"])
for pool, n in pool_cnt.most_common(20):
    ags = [d["ag"] for d in BUE if d["pool"] == pool and d["ag"]]
    vms = [d["vm"] for d in BUE if d["pool"] == pool and d["vm"]]
    m_ag = collections.Counter(maske(v) for v in ags).most_common(1)
    m_vm = collections.Counter(maske(v) for v in vms).most_common(1)
    setz(zeile, 1, pool, sz=10, wrap=True, rahmen=True)
    setz(zeile, 2, n, sz=10, fmt="#,##0", rahmen=True)
    setz(zeile, 3, "%s (%dx)" % (m_ag[0][0], m_ag[0][1]) if m_ag else "-", sz=10, rahmen=True)
    setz(zeile, 4, next((v for v in ags if maske(v) == m_ag[0][0]), "") if m_ag else "",
         sz=10, rahmen=True)
    setz(zeile, 5, "%s (%dx)" % (m_vm[0][0], m_vm[0][1]) if m_vm else "-", sz=10, rahmen=True)
    setz(zeile, 6, next((v for v in vms if maske(v) == m_vm[0][0]), "") if m_vm else "",
         sz=10, rahmen=True)
    setz(zeile, 7, "%d / %d" % (len({maske(v) for v in ags}), len({maske(v) for v in vms})),
         sz=10, rahmen=True)
    zeile += 1
zeile += 1


# ---- Block 9: Legende ---------------------------------------------------
titel("9 - Lesehilfe", "")
LEGENDE = [
    ("Muster-Schreibweise", "9 = eine Ziffer, A = ein Großbuchstabe, a = ein Kleinbuchstabe. "
                            "'9x7' = sieben Ziffern hintereinander. Trennzeichen stehen "
                            "unverändert im Muster: '9x5-9x4' = fünf Ziffern, Bindestrich, "
                            "vier Ziffern."),
    ("Basis", "Ausgewertet werden ausschließlich BÜ-Vorgänge, also Zeilen mit "
              "Klassifikation '%s' (%s von %s). Alles andere - Antrag/Änderung, "
              "Anfrage/Rücksprache, Kein-Makler-Vorgang, ohne Klassifikation - ist "
              "ausgeschlossen." % (BUE_WERT, zahl(N_BUE), zahl(N))),
    ("Anzahl-Spalten", "COUNTIFS-Formeln auf das Hilfsblatt '%s', immer mit der Bedingung "
                       "BUe_Vorgang='ja'. Wer dort filtert oder Zeilen ergänzt, sieht die "
                       "Auswertung sofort mitlaufen." % SHEET_DET),
    ("Farben", "Gelb = Einzelfall oder uneinheitliche Schreibweise. Rot = formal fehlerhafte "
               "Extraktion. Grün = Prüfung ohne Befund."),
    ("Hilfsblatt", "'%s' enthält ALLE %s Zeilen mit Rohwert, Muster, Klartext-Beschreibung, "
                   "Ziffernanzahl und Trennzeichen. Die Spalte BUe_Vorgang zeigt, welche "
                   "Zeilen in die Auswertung eingehen." % (SHEET_DET, zahl(N))),
]
for a, b in LEGENDE:
    setz(zeile, 1, a, bold=True, sz=10, wrap=True, rahmen=True, fill=GRAU)
    setz(zeile, 2, b, sz=10, wrap=True, rahmen=True)
    for c in range(3, 8):
        setz(zeile, c, None, rahmen=True)
    ms.merge_cells(start_row=zeile, start_column=2, end_row=zeile, end_column=7)
    ms.row_dimensions[zeile].height = 14 * max(2, math.ceil(len(b) / 130.0))
    zeile += 1

ms.freeze_panes = "A5"
ms.sheet_view.showGridLines = False

# Excel soll die COUNTIF-Formeln beim Oeffnen einmal komplett durchrechnen
# (openpyxl schreibt Formeln ohne zwischengespeicherte Ergebnisse).
wb.calculation.fullCalcOnLoad = True

wb.save(DATEI)
print("OK - '%s' und '%s' geschrieben in %s (%d Datenzeilen)" % (SHEET_MAIN, SHEET_DET, DATEI, N))
