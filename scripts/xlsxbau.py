#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gemeinsame Bausteine fuer die Excel-Anlagen in diesem Ordner.

Damit sehen alle Mappen gleich aus: Titel in Zeile 1, Erlaeuterung darunter,
Kopfzeile mit Farbbalken, umrandete Datenzeilen, Autofilter und fixierte
Kopfzeile, keine Gitternetzlinien, Schrift Aptos.
"""

import math

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

FONT = "Aptos"

BLAU = "FF1E40AF"
LILA = "FF4C1D95"
GRAU = "FFF3F4F6"
GELB = "FFFEF3C7"
ROT = "FFFEE2E2"
GRUEN = "FFDCFCE7"
WEISS = "FFFFFFFF"
TEXTGRAU = "FF4B5563"

THIN = Side(style="thin", color="FFD1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def titel(ws, text, *erlaeuterungen):
    """Titelzeile plus beliebig viele Erlaeuterungszeilen darunter."""
    ws["A1"] = text
    ws["A1"].font = Font(name=FONT, sz=14, b=True, color=BLAU)
    for i, zeile in enumerate(erlaeuterungen, 2):
        ws.cell(row=i, column=1, value=zeile).font = Font(name=FONT, sz=10, color=TEXTGRAU)
    return len(erlaeuterungen) + 2      # naechste freie Zeile


def kopfzeile(ws, spalten, breiten, zeile):
    for i, (t, b) in enumerate(zip(spalten, breiten), 1):
        z = ws.cell(row=zeile, column=i, value=t)
        z.font = Font(name=FONT, sz=10, b=True, color=WEISS)
        z.fill = PatternFill("solid", fgColor=LILA)
        z.alignment = Alignment(vertical="center", wrap_text=True)
        z.border = BORDER
        ws.column_dimensions[get_column_letter(i)].width = b
    ws.row_dimensions[zeile].height = 24
    return zeile + 1


def schreibe(ws, zeile, werte, breiten=None, fill=None, fett_spalten=()):
    """Datenzeile mit Rahmen, Umbruch und - falls breiten gesetzt - Zeilenhoehe."""
    for i, wert in enumerate(werte, 1):
        z = ws.cell(row=zeile, column=i, value=wert)
        z.font = Font(name=FONT, sz=10, b=(i in fett_spalten))
        z.alignment = Alignment(vertical="top", wrap_text=True)
        z.border = BORDER
        if fill:
            z.fill = PatternFill("solid", fgColor=fill)
    if breiten:
        noetig = [math.ceil(len(w) / max(10, b * 1.05))
                  for w, b in zip(werte, breiten) if isinstance(w, str) and w]
        ws.row_dimensions[zeile].height = 14 * min(max(noetig or [1], default=1), 8)
    return zeile + 1


def hinweis(ws, zeile, text, spalten, farbe="FF92400E"):
    """Fussnote unter einer Tabelle, ueber die Tabellenbreite verbunden."""
    z = ws.cell(row=zeile, column=1, value=text)
    z.font = Font(name=FONT, sz=9, color=farbe)
    z.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=zeile, start_column=1, end_row=zeile, end_column=spalten)
    ws.row_dimensions[zeile].height = 14 * max(2, math.ceil(len(text) / 130.0))
    return zeile + 1


def abschluss(ws, kopf_zeile, letzte_zeile, spalten, fixieren=None):
    """Autofilter, fixierte Kopfzeile und Gitternetz aus."""
    if letzte_zeile > kopf_zeile:
        ws.auto_filter.ref = "A%d:%s%d" % (kopf_zeile, get_column_letter(spalten),
                                           letzte_zeile)
    ws.freeze_panes = fixieren or ("A%d" % (kopf_zeile + 1))
    ws.sheet_view.showGridLines = False
