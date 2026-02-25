#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
BGAV Titelbezeichnung – Test-Mail-Extraktion
=============================================

Erweitert das bestehende E-Mail-Tool um eine Python-basierte Extraktion
von gesendeten E-Mails aus dem Exchange-Postfach umstellung1ev@barmenia.de.

Funktionen:
- Verbindung zum Exchange-Server via EWS (exchangelib)
- Suche im Gesendete-Elemente-Ordner nach "BGAV | Titelbezeichnung"
- Export als .eml und .pdf
- Metadaten-Abgleich mit Übersicht_Titeländerung.xlsx
- Erzeugung einer testfaelle_metadaten.csv

Verwendung:
    # Umgebungsvariablen setzen:
    export EWS_EMAIL=umstellung1ev@barmenia.de
    export EWS_PASSWORD=<passwort>
    # Optional:
    export EWS_SERVER=<exchange-server>  # Falls Autodiscover nicht funktioniert
    export EWS_MAX_MAILS=5               # Anzahl der zu extrahierenden Mails (Standard: 5)

    python3 extract_bgav_testmails.py

Abhängigkeiten:
    pip install exchangelib openpyxl xhtml2pdf
"""

import os
import sys
import csv
import email
import logging
from email import policy
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formataddr, formatdate
from datetime import datetime
from pathlib import Path
from io import BytesIO

import openpyxl

# ─── Logging ─────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
log = logging.getLogger(__name__)

# ─── Konfiguration ───────────────────────────────────────────────────
MAILBOX_EMAIL = os.environ.get('EWS_EMAIL', 'umstellung1ev@barmenia.de')
MAILBOX_PASSWORD = os.environ.get('EWS_PASSWORD', '')
EWS_SERVER = os.environ.get('EWS_SERVER', '')
MAX_MAILS = int(os.environ.get('EWS_MAX_MAILS', '5'))

SUBJECT_FILTER = 'BGAV | Titelbezeichnung'
KLAMMERBEGRIFF = 'BGAV Titelbezeichnung'

SCRIPT_DIR = Path(__file__).resolve().parent
EXCEL_FILE = SCRIPT_DIR / 'Übersicht_Titeländerung.xlsx'
OUTPUT_DIR = SCRIPT_DIR / 'output'
OUTPUT_EML = OUTPUT_DIR / 'eml'
OUTPUT_PDF = OUTPUT_DIR / 'pdf'

FILE_PREFIX = 'BGAV_Titelbezeichnung'


# ─── Excel-Metadaten laden ───────────────────────────────────────────

def load_excel_metadata(excel_path: Path) -> dict:
    """
    Liest die Übersicht_Titeländerung.xlsx und erstellt ein Lookup-Dict
    von E-Mail-Adresse (lowercase) -> {bd_nummer, vorname, nachname, email}.

    Spalten:
      A = BD-Nummer (= VD-Wechsler-Neue-Nummer im Header, aber Spalte A enthält "Match")
      I = Vorname
      J = Nachname
      K = Email
    """
    log.info(f'Lade Excel-Datei: {excel_path}')
    if not excel_path.exists():
        log.error(f'Excel-Datei nicht gefunden: {excel_path}')
        sys.exit(1)

    wb = openpyxl.load_workbook(str(excel_path), data_only=True)
    ws = wb.active
    log.info(f'Sheet: "{ws.title}", Zeilen: {ws.max_row}')

    lookup = {}
    for row in range(2, ws.max_row + 1):
        bd_nummer = ws.cell(row=row, column=1).value  # Spalte A
        vorname = ws.cell(row=row, column=9).value     # Spalte I
        nachname = ws.cell(row=row, column=10).value   # Spalte J
        email_addr = ws.cell(row=row, column=11).value  # Spalte K

        if email_addr:
            key = str(email_addr).strip().lower()
            lookup[key] = {
                'bd_nummer': str(bd_nummer).strip() if bd_nummer else '',
                'vorname': str(vorname).strip() if vorname else '',
                'nachname': str(nachname).strip() if nachname else '',
                'email': str(email_addr).strip(),
            }

    wb.close()
    log.info(f'{len(lookup)} Einträge aus Excel geladen')
    return lookup


# ─── Exchange-Verbindung ─────────────────────────────────────────────

def connect_exchange():
    """
    Stellt eine Verbindung zum Exchange-Server via EWS her.
    Verwendet Autodiscover oder explizite Server-Angabe.
    """
    from exchangelib import (
        Credentials, Account, Configuration, DELEGATE,
        Build, Version
    )

    if not MAILBOX_PASSWORD:
        log.error(
            'EWS_PASSWORD nicht gesetzt. Bitte setzen:\n'
            '  export EWS_PASSWORD=<passwort>\n'
            '  export EWS_EMAIL=umstellung1ev@barmenia.de'
        )
        sys.exit(1)

    log.info(f'Verbinde mit Exchange-Postfach: {MAILBOX_EMAIL}')
    credentials = Credentials(MAILBOX_EMAIL, MAILBOX_PASSWORD)

    if EWS_SERVER:
        log.info(f'Verwende expliziten Server: {EWS_SERVER}')
        config = Configuration(
            server=EWS_SERVER,
            credentials=credentials,
        )
        account = Account(
            primary_smtp_address=MAILBOX_EMAIL,
            config=config,
            autodiscover=False,
            access_type=DELEGATE,
        )
    else:
        log.info('Verwende Autodiscover...')
        account = Account(
            primary_smtp_address=MAILBOX_EMAIL,
            credentials=credentials,
            autodiscover=True,
            access_type=DELEGATE,
        )

    log.info(f'Verbunden mit Postfach: {account.primary_smtp_address}')
    return account


# ─── E-Mails suchen und extrahieren ─────────────────────────────────

def search_sent_items(account, subject_filter: str, max_results: int):
    """
    Durchsucht den Gesendete-Elemente-Ordner nach Mails mit dem
    angegebenen Betreff-Filter.

    Gibt nur gesendete Mails zurück (FROM = Postfach-Adresse).
    """
    from exchangelib import Q

    sent_folder = account.sent

    log.info(f'Durchsuche Gesendete Elemente nach: "{subject_filter}"')
    log.info(f'Gesendete-Ordner: {sent_folder.name} ({sent_folder.total_count} Elemente)')

    # Filtere nach Betreff (enthält den Suchbegriff)
    query = sent_folder.filter(
        subject__contains=subject_filter
    ).order_by('-datetime_sent')

    results = []
    for item in query:
        # Sicherheitscheck: Nur Mails die tatsächlich vom Postfach gesendet wurden
        sender_email = ''
        if item.sender and item.sender.email_address:
            sender_email = item.sender.email_address.lower()

        if sender_email != MAILBOX_EMAIL.lower():
            log.warning(
                f'  Überspringe Mail (Absender {sender_email} != {MAILBOX_EMAIL}): '
                f'{item.subject}'
            )
            continue

        results.append(item)
        recipients = []
        if item.to_recipients:
            recipients = [r.email_address for r in item.to_recipients if r.email_address]

        log.info(
            f'  [{len(results)}] {item.subject} | '
            f'An: {", ".join(recipients)} | '
            f'Gesendet: {item.datetime_sent}'
        )

        if len(results) >= max_results:
            break

    log.info(f'{len(results)} Mails gefunden (von max. {max_results})')
    return results


# ─── EML-Export ──────────────────────────────────────────────────────

def save_as_eml(item, filepath: Path):
    """
    Speichert eine Exchange-Mail als .eml-Datei (RFC 2822).
    Verwendet die MIME-Rohdaten der Mail wenn verfügbar,
    andernfalls wird die EML aus den Metadaten zusammengebaut.
    """
    try:
        # Versuche die MIME-Rohdaten direkt zu bekommen
        mime_content = item.mime_content
        if mime_content:
            with open(filepath, 'wb') as f:
                f.write(mime_content)
            log.info(f'  EML gespeichert (MIME): {filepath.name}')
            return
    except Exception as e:
        log.warning(f'  MIME-Content nicht verfügbar: {e}')

    # Fallback: EML aus Metadaten zusammenbauen
    msg = MIMEMultipart('alternative')

    # Header setzen
    msg['Subject'] = item.subject or ''

    sender_name = ''
    sender_email = MAILBOX_EMAIL
    if item.sender:
        sender_name = item.sender.name or ''
        sender_email = item.sender.email_address or MAILBOX_EMAIL
    msg['From'] = formataddr((sender_name, sender_email))

    if item.to_recipients:
        to_addrs = []
        for r in item.to_recipients:
            to_addrs.append(formataddr((r.name or '', r.email_address or '')))
        msg['To'] = ', '.join(to_addrs)

    if item.datetime_sent:
        msg['Date'] = formatdate(item.datetime_sent.timestamp(), localtime=True)

    if item.message_id:
        msg['Message-ID'] = item.message_id

    # Body hinzufügen
    if item.body:
        body_text = str(item.body)
        if item.body.body_type == 'HTML':
            msg.attach(MIMEText(body_text, 'html', 'utf-8'))
        else:
            msg.attach(MIMEText(body_text, 'plain', 'utf-8'))

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(msg.as_string())

    log.info(f'  EML gespeichert (rekonstruiert): {filepath.name}')


# ─── PDF-Export ──────────────────────────────────────────────────────

def save_as_pdf(item, filepath: Path):
    """
    Rendert eine E-Mail als PDF.
    Erstellt ein HTML-Layout mit Header-Informationen und Body-Inhalt.
    """
    from xhtml2pdf import pisa

    # Empfänger aufbereiten
    recipients = []
    if item.to_recipients:
        for r in item.to_recipients:
            name = r.name or ''
            addr = r.email_address or ''
            if name and addr:
                recipients.append(f'{name} &lt;{addr}&gt;')
            else:
                recipients.append(addr or name)

    cc_list = []
    if item.cc_recipients:
        for r in item.cc_recipients:
            name = r.name or ''
            addr = r.email_address or ''
            if name and addr:
                cc_list.append(f'{name} &lt;{addr}&gt;')
            else:
                cc_list.append(addr or name)

    # Datum formatieren
    sent_date = ''
    if item.datetime_sent:
        sent_date = item.datetime_sent.strftime('%d.%m.%Y %H:%M')

    # Sender
    sender_display = MAILBOX_EMAIL
    if item.sender:
        if item.sender.name:
            sender_display = f'{item.sender.name} &lt;{item.sender.email_address}&gt;'
        elif item.sender.email_address:
            sender_display = item.sender.email_address

    # Body aufbereiten
    body_html = ''
    if item.body:
        body_text = str(item.body)
        if item.body.body_type == 'HTML':
            # HTML-Body: Inline-Styles und externe Referenzen entfernen
            body_html = _clean_html_for_pdf(body_text)
        else:
            # Plaintext in HTML konvertieren
            body_html = f'<pre style="white-space:pre-wrap;font-family:Arial,sans-serif;font-size:10pt;">{_escape_html(body_text)}</pre>'

    # PDF-HTML zusammenbauen
    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8"/>
    <style>
        @page {{
            size: A4;
            margin: 2cm;
        }}
        body {{
            font-family: Arial, Helvetica, sans-serif;
            font-size: 10pt;
            color: #333;
        }}
        .email-header {{
            border-bottom: 2px solid #0078d4;
            padding-bottom: 10px;
            margin-bottom: 15px;
        }}
        .header-field {{
            margin: 3px 0;
            font-size: 9pt;
        }}
        .header-label {{
            font-weight: bold;
            color: #555;
            display: inline-block;
            width: 60px;
        }}
        .subject {{
            font-size: 14pt;
            font-weight: bold;
            color: #0078d4;
            margin: 10px 0 5px 0;
        }}
        .email-body {{
            margin-top: 10px;
            line-height: 1.4;
        }}
        .footer {{
            margin-top: 20px;
            padding-top: 10px;
            border-top: 1px solid #ccc;
            font-size: 8pt;
            color: #999;
        }}
    </style>
</head>
<body>
    <div class="email-header">
        <div class="subject">{_escape_html(item.subject or '')}</div>
        <div class="header-field">
            <span class="header-label">Von:</span> {sender_display}
        </div>
        <div class="header-field">
            <span class="header-label">An:</span> {'; '.join(recipients)}
        </div>
        {'<div class="header-field"><span class="header-label">CC:</span> ' + '; '.join(cc_list) + '</div>' if cc_list else ''}
        <div class="header-field">
            <span class="header-label">Datum:</span> {sent_date}
        </div>
        <div class="header-field">
            <span class="header-label">Betreff:</span> {_escape_html(item.subject or '')}
        </div>
    </div>
    <div class="email-body">
        {body_html}
    </div>
    <div class="footer">
        Exportiert am {datetime.now().strftime('%d.%m.%Y %H:%M')} |
        Quelle: {MAILBOX_EMAIL} (Gesendete Elemente)
    </div>
</body>
</html>"""

    # PDF erzeugen
    result_file = BytesIO()
    pisa_status = pisa.CreatePDF(html_content, dest=result_file, encoding='utf-8')

    if pisa_status.err:
        log.warning(f'  PDF-Rendering hatte Fehler (wird trotzdem gespeichert): {filepath.name}')

    with open(filepath, 'wb') as f:
        f.write(result_file.getvalue())

    log.info(f'  PDF gespeichert: {filepath.name}')


def _escape_html(text: str) -> str:
    """Escaped HTML-Sonderzeichen."""
    return (text
        .replace('&', '&amp;')
        .replace('<', '&lt;')
        .replace('>', '&gt;')
        .replace('"', '&quot;')
    )


def _clean_html_for_pdf(html: str) -> str:
    """
    Bereinigt HTML-Body für die PDF-Erzeugung.
    Entfernt externe Referenzen und problematische Elemente.
    """
    import re

    # Entferne <style>-Blöcke (können Probleme verursachen)
    html = re.sub(r'<style[^>]*>.*?</style>', '', html, flags=re.DOTALL | re.IGNORECASE)

    # Entferne <script>-Blöcke
    html = re.sub(r'<script[^>]*>.*?</script>', '', html, flags=re.DOTALL | re.IGNORECASE)

    # Entferne <link>-Tags
    html = re.sub(r'<link[^>]*/?>', '', html, flags=re.IGNORECASE)

    # Entferne <meta>-Tags
    html = re.sub(r'<meta[^>]*/?>', '', html, flags=re.IGNORECASE)

    # Entferne <html>, <head>, <body> Tags (wir haben unsere eigenen)
    html = re.sub(r'</?html[^>]*>', '', html, flags=re.IGNORECASE)
    html = re.sub(r'</?head[^>]*>', '', html, flags=re.IGNORECASE)
    html = re.sub(r'</?body[^>]*>', '', html, flags=re.IGNORECASE)
    html = re.sub(r'<!DOCTYPE[^>]*>', '', html, flags=re.IGNORECASE)

    # Entferne <img>-Tags mit externen src (CID-Bilder auch, da nicht eingebettet)
    html = re.sub(r'<img[^>]*>', '', html, flags=re.IGNORECASE)

    return html.strip()


# ─── Metadaten-CSV erzeugen ──────────────────────────────────────────

def generate_metadata_csv(
    mail_items: list,
    excel_lookup: dict,
    output_path: Path
):
    """
    Erzeugt die testfaelle_metadaten.csv mit allen Metadaten.

    Spalten: Nr;Dateiname_EML;Dateiname_PDF;Empfaenger_Email;Vorname;Nachname;BD_Nummer;Klammerbegriff

    Falls eine Mail mehrere Empfänger hat, wird pro Empfänger ein Eintrag erzeugt.
    """
    rows = []
    nr = 0

    for idx, item in enumerate(mail_items):
        file_nr = f'{idx + 1:03d}'
        eml_name = f'{FILE_PREFIX}_{file_nr}.eml'
        pdf_name = f'{FILE_PREFIX}_{file_nr}.pdf'

        recipients = []
        if item.to_recipients:
            recipients = [r for r in item.to_recipients if r.email_address]

        if not recipients:
            # Kein Empfänger gefunden – trotzdem einen Eintrag erzeugen
            nr += 1
            rows.append({
                'Nr': nr,
                'Dateiname_EML': eml_name,
                'Dateiname_PDF': pdf_name,
                'Empfaenger_Email': '',
                'Vorname': '',
                'Nachname': '',
                'BD_Nummer': '',
                'Klammerbegriff': KLAMMERBEGRIFF,
            })
            log.warning(f'  Mail {file_nr}: Kein Empfänger gefunden')
            continue

        for recipient in recipients:
            nr += 1
            rcpt_email = recipient.email_address.strip().lower()
            meta = excel_lookup.get(rcpt_email, None)

            if meta:
                rows.append({
                    'Nr': nr,
                    'Dateiname_EML': eml_name,
                    'Dateiname_PDF': pdf_name,
                    'Empfaenger_Email': meta['email'],
                    'Vorname': meta['vorname'],
                    'Nachname': meta['nachname'],
                    'BD_Nummer': meta['bd_nummer'],
                    'Klammerbegriff': KLAMMERBEGRIFF,
                })
                log.info(
                    f'  Mail {file_nr} -> {meta["email"]} | '
                    f'BD: {meta["bd_nummer"]} | '
                    f'{meta["vorname"]} {meta["nachname"]}'
                )
            else:
                # Empfänger nicht in Excel gefunden
                rows.append({
                    'Nr': nr,
                    'Dateiname_EML': eml_name,
                    'Dateiname_PDF': pdf_name,
                    'Empfaenger_Email': recipient.email_address,
                    'Vorname': recipient.name or '',
                    'Nachname': '',
                    'BD_Nummer': '',
                    'Klammerbegriff': KLAMMERBEGRIFF,
                })
                log.warning(
                    f'  Mail {file_nr} -> {recipient.email_address}: '
                    f'NICHT in Excel gefunden! BD-Nummer leer.'
                )

    # CSV schreiben (UTF-8, Semikolon-getrennt)
    fieldnames = [
        'Nr', 'Dateiname_EML', 'Dateiname_PDF',
        'Empfaenger_Email', 'Vorname', 'Nachname',
        'BD_Nummer', 'Klammerbegriff'
    ]

    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
        writer.writeheader()
        writer.writerows(rows)

    log.info(f'Metadaten-CSV gespeichert: {output_path} ({len(rows)} Einträge)')
    return rows


# ─── Hauptprogramm ───────────────────────────────────────────────────

def main():
    log.info('=' * 60)
    log.info('BGAV Titelbezeichnung – Test-Mail-Extraktion')
    log.info('=' * 60)

    # 1. Ausgabeverzeichnisse erstellen
    OUTPUT_EML.mkdir(parents=True, exist_ok=True)
    OUTPUT_PDF.mkdir(parents=True, exist_ok=True)
    log.info(f'Ausgabeverzeichnis: {OUTPUT_DIR}')

    # 2. Excel-Metadaten laden
    excel_lookup = load_excel_metadata(EXCEL_FILE)

    # 3. Exchange-Verbindung herstellen
    account = connect_exchange()

    # 4. Gesendete Mails suchen
    mail_items = search_sent_items(account, SUBJECT_FILTER, MAX_MAILS)

    if not mail_items:
        log.error(
            f'Keine Mails mit Betreff "{SUBJECT_FILTER}" im '
            f'Gesendete-Elemente-Ordner gefunden!'
        )
        sys.exit(1)

    if len(mail_items) < MAX_MAILS:
        log.warning(
            f'Nur {len(mail_items)} von {MAX_MAILS} gewünschten Mails gefunden'
        )

    # 5. Mails exportieren (EML + PDF)
    log.info('')
    log.info('Exportiere Mails...')
    for idx, item in enumerate(mail_items):
        file_nr = f'{idx + 1:03d}'
        eml_path = OUTPUT_EML / f'{FILE_PREFIX}_{file_nr}.eml'
        pdf_path = OUTPUT_PDF / f'{FILE_PREFIX}_{file_nr}.pdf'

        log.info(f'Mail {file_nr}: {item.subject}')
        save_as_eml(item, eml_path)
        save_as_pdf(item, pdf_path)

    # 6. Metadaten-CSV erzeugen
    log.info('')
    log.info('Erzeuge Metadaten-CSV...')
    csv_path = OUTPUT_DIR / 'testfaelle_metadaten.csv'
    rows = generate_metadata_csv(mail_items, excel_lookup, csv_path)

    # 7. Zusammenfassung
    log.info('')
    log.info('=' * 60)
    log.info('ERGEBNIS')
    log.info('=' * 60)
    log.info(f'EML-Dateien:  {len(mail_items)} in {OUTPUT_EML}')
    log.info(f'PDF-Dateien:  {len(mail_items)} in {OUTPUT_PDF}')
    log.info(f'CSV-Datei:    {csv_path}')
    log.info(f'CSV-Einträge: {len(rows)}')
    log.info('')

    # Dateiübersicht ausgeben
    log.info('Dateien:')
    for d in sorted(OUTPUT_EML.glob('*.eml')):
        log.info(f'  eml/{d.name}')
    for d in sorted(OUTPUT_PDF.glob('*.pdf')):
        log.info(f'  pdf/{d.name}')
    log.info(f'  {csv_path.name}')

    log.info('')
    log.info('Fertig! Die Dateien können als Input für das Java-Batch-Tool verwendet werden.')


if __name__ == '__main__':
    main()
